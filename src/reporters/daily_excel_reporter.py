"""
Enhanced Daily Excel Reporter with Professional Formatting and Pivot Tables
"""
from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import logging

from config.settings import settings

logger = logging.getLogger(__name__)

class DailyExcelReporter:
    """Enhanced daily reporter with professional formatting and pivot analysis"""
    
    def __init__(self):
        self.output_dir = settings.report_output_dir
        self.colors = {
            'header_bg': '1E2761',      # Dark blue
            'header_text': 'FFFFFF',     # White
            'critical': 'DC143C',        # Crimson red
            'high': 'FF8C00',            # Dark orange
            'medium': 'FFD700',          # Gold
            'low': '90EE90',             # Light green
            'alt_row': 'F8F9FA'          # Light gray
        }
    
    def generate_daily_report(self, summary: Dict[str, Any]) -> Path:
        """
        Generate professional daily Excel report

        Args:
            summary: Dictionary containing critical_items and exposed_secrets

        Returns:
            Path to generated report
        """
        from src.utils.validators import DataValidator
        org_name  = summary.get('organization', settings.github_org)
        sanitized = DataValidator.sanitize_filename(org_name)
        ts        = datetime.now().strftime('%Y%m%d_%H%M%S')

        filename  = self.output_dir / f'{sanitized}_Daily_Report_{ts}.xlsx'

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Tab 1 – Executive Summary (must be created first so it is leftmost)
            self._sheet_executive_summary(writer, summary)

            # Tab 2 – Critical Items
            if summary.get('critical_items'):
                self._create_critical_items_sheet(writer, summary['critical_items'])

            # Tab 3 – Exposed Secrets
            if summary.get('exposed_secrets'):
                self._create_exposed_secrets_sheet(writer, summary['exposed_secrets'])



        # Apply professional formatting (Executive Summary is skipped — already styled)
        self._apply_formatting(filename, summary)

        logger.info(f"[bright_green]✓ Daily report generated successfully[/bright_green]")
        return filename
    

    # ── Tab 1 – Executive Summary ─────────────────────────────────────────────
    def _sheet_executive_summary(self, writer, summary: Dict[str, Any]):
        """
        Build the Executive Summary tab — identical layout to the weekly report.

        Sections
        --------
        1. Banner           – tool title, report type, generated date, period
        2. Org Details      – org name, contact email, GitHub URL
        3. Security Summary – new alert counts by severity with action required
        4. About This Tool  – version, developer, website, support contact
        """
        # ── Safe settings reader ──────────────────────────────────────────────
        def _s(attr: str, default: str = '') -> str:
            val = getattr(settings, attr, None)
            if val is None:
                return default
            if hasattr(val, 'default'):
                fb = val.default
                return str(fb) if fb is not None else default
            raw = str(val).strip()
            return raw if raw else default

        # ── Pull metadata from settings (.env) ────────────────────────────────
        tool_title   = _s('report_title',     'GitHub Advanced Security Reporter')
        company      = _s('company_name',     '') or settings.github_org
        developed_by = _s('developed_by',     'Security Engineering Team')
        version      = _s('tool_version',     '1.0.0')
        copy_year    = _s('copyright_year',   '') or str(datetime.now().year)
        website      = _s('company_website',  '') or 'N/A'
        github_repo  = _s('tool_github_repo', '') or 'N/A'
        support_mail = _s('support_email',    '') or 'N/A'

        # ── Alert counts from summary ─────────────────────────────────────────
        items         = summary.get('critical_items', [])
        secrets_count = len(summary.get('exposed_secrets', []))

        sev  = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
        type_counts: Dict[str, int] = {}
        for item in items:
            s = str(item.get('severity') or '').lower()
            if s in sev:
                sev[s] += 1
            t = item.get('type', 'Unknown')
            type_counts[t] = type_counts.get(t, 0) + 1
        total_vulns = sum(sev.values())

        # ── Colour / style helpers (self-contained, no shared globals needed) ─
        CLR = {
            'navy':    '1E2761', 'white':  'FFFFFF', 'teal':   '065A82',
            'green':   '2C5F2D', 'dkgreen':'006400', 'red':    'DC143C',
            'dkred':   '8B0000', 'orange': 'FF8C00', 'gold':   'FFD700',
            'lime':    '90EE90', 'pink':   'FFB6C1', 'alt':    'F8F9FA',
        }
        _thin   = Side(style='thin')
        _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        def _fill(hex_: str) -> PatternFill:
            return PatternFill(start_color=hex_, end_color=hex_, fill_type='solid')

        def _sev_style(sev_key: str):
            if sev_key == 'critical': return CLR['red'],    CLR['white']
            if sev_key == 'high':     return CLR['orange'], CLR['white']
            if sev_key == 'medium':   return CLR['gold'],   '000000'
            if sev_key == 'low':      return CLR['lime'],   '000000'
            return None, None

        def _merged_hdr(ws, row: int, text: str, bg: str,
                        ncols: int = 4, size: int = 13, height: int = 22):
            cell = ws.cell(row=row, column=1, value=text)
            ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
            cell.font      = Font(bold=True, color=CLR['white'], size=size)
            cell.fill      = _fill(bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = height

        def _col_hdr(ws, row: int, labels: list, bg: str = CLR['navy']):
            for c, label in enumerate(labels, 1):
                cell           = ws.cell(row=row, column=c, value=label)
                cell.font      = Font(bold=True, color=CLR['white'], size=11)
                cell.fill      = _fill(bg)
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
                cell.border    = _border
            ws.row_dimensions[row].height = 18

        # ── Create blank worksheet ────────────────────────────────────────────
        pd.DataFrame({'_': []}).to_excel(
            writer, sheet_name='Executive Summary', index=False)
        ws = writer.sheets['Executive Summary']
        ws.delete_rows(1)          # remove the dummy header row

        r = 1   # row pointer

        # ══════════════════════════════════════════════════════════════════════
        #  SECTION 1 – Banner
        # ══════════════════════════════════════════════════════════════════════
        _merged_hdr(ws, r, f'🛡  {tool_title}', CLR['navy'], size=15, height=28)
        r += 1
        for label, value in [
            ('Report Type',   'Daily Security Report'),
            ('Generated On',  datetime.now().strftime('%A, %d %B %Y  —  %H:%M:%S')),
            ('Report Period', 'Daily – Last 24 hours'),
        ]:
            ws.cell(r, 1, label).font = Font(bold=True, color=CLR['navy'], size=11)
            ws.cell(r, 2, value).font = Font(size=11)
            r += 1
        r += 1

        # ══════════════════════════════════════════════════════════════════════
        #  SECTION 2 – Organisation Details
        # ══════════════════════════════════════════════════════════════════════
        _merged_hdr(ws, r, '🏢  Organization Details', CLR['green'], size=12)
        r += 1
        for label, value in [
            ('Organization Name',   company),
            ('GitHub Organization', settings.github_org),
            ('Contact / Support',   support_mail),
            ('GitHub URL',          f'https://github.com/{settings.github_org}'),
        ]:
            lc = ws.cell(r, 1, label)
            vc = ws.cell(r, 2, value)
            lc.font = Font(bold=True, color='444444', size=11)
            vc.font = Font(size=11)
            if str(value).startswith('http') or ('@' in str(value) and '.' in str(value)):
                vc.font = Font(size=11, color='0563C1', underline='single')
            if r % 2 == 0:
                lc.fill = _fill(CLR['alt'])
                vc.fill = _fill(CLR['alt'])
            r += 1
        r += 1

        # ══════════════════════════════════════════════════════════════════════
        #  SECTION 3 – Security Summary (New Alerts — Last 24 Hours)
        # ══════════════════════════════════════════════════════════════════════
        _merged_hdr(ws, r, '⚠️  Security Summary', CLR['red'], size=12)
        r += 1
        _col_hdr(ws, r, ['Metric', 'Count', 'Status', 'Action Required'], CLR['navy'])
        r += 1

        sev_rows = [
            ('Total New Alerts Today',
             total_vulns, '',
             '⚠ Review Now'       if total_vulns          > 0 else '✓ None',
             'Triage all new alerts'),
            ('  ↳ Critical Severity',
             sev['critical'], 'critical',
             '🔴 Fix Immediately' if sev['critical']       > 0 else '✓ None',
             'Patch within 48 hrs'),
            ('  ↳ High Severity',
             sev['high'], 'high',
             '🟠 Fix This Week'   if sev['high']           > 0 else '✓ None',
             'Patch within 7 days'),
            ('  ↳ Medium Severity',
             sev['medium'], 'medium',
             '🟡 Plan Fix'        if sev['medium']         > 0 else '✓ None',
             'Patch within 30 days'),
            ('  ↳ Low Severity',
             sev['low'], 'low',
             '🟢 Monitor'         if sev['low']            > 0 else '✓ None',
             'Patch within 90 days'),
            ('Exposed Secrets',
             secrets_count, 'critical',
             '🔴 Fix Immediately' if secrets_count         > 0 else '✓ None',
             'Revoke & rotate NOW'),
        ]
        for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
            sev_rows.append((f'  ↳ {t}', cnt, '', '', 'See detail tab'))

        for label, count, sev_key, status, action in sev_rows:
            mc = ws.cell(r, 1, label);  cc = ws.cell(r, 2, count)
            sc = ws.cell(r, 3, status); nc = ws.cell(r, 4, action)
            mc.font      = Font(bold=not label.startswith('  '), size=11)
            cc.alignment = Alignment(horizontal='center')
            nc.font      = Font(italic=True, color='555555', size=10)
            bg, fg = _sev_style(sev_key)
            if bg and count > 0:
                cc.fill = _fill(bg);  cc.font = Font(bold=True, color=fg, size=11)
            if   '🔴' in status: sc.font = Font(bold=True, color=CLR['dkred'],  size=11)
            elif '🟠' in status: sc.font = Font(bold=True, color='8B4513',       size=11)
            elif '🟡' in status: sc.font = Font(bold=True, color='8B8B00',       size=11)
            elif '✓'  in status: sc.font = Font(bold=True, color=CLR['dkgreen'], size=11)
            for cell in (mc, cc, sc, nc):
                cell.border = _border
                if r % 2 == 0:
                    try:
                        if cell.fill.start_color.rgb in ('00000000', ''):
                            cell.fill = _fill(CLR['alt'])
                    except Exception:
                        pass
            r += 1
        r += 1

        # ══════════════════════════════════════════════════════════════════════
        #  SECTION 4 – About This Tool
        # ══════════════════════════════════════════════════════════════════════
        _merged_hdr(ws, r, '📋  About This Tool', CLR['teal'], size=12)
        r += 1
        for label, value in [
            ('Tool Name',         tool_title),
            ('Company / Client',  company),
            ('Developed By',      developed_by),
            ('Version',           f'v{version}'),
            ('Copyright',         f'© {copy_year}  {company}'),
            ('Company Website',   website),
            ('GitHub Repository', github_repo),
            ('Support / Contact', support_mail),
        ]:
            lc = ws.cell(r, 1, label);  vc = ws.cell(r, 2, value)
            lc.font = Font(bold=True, color='444444', size=11)
            vc.font = Font(size=11)
            if str(value).startswith('http') or ('@' in str(value) and '.' in str(value)):
                vc.font = Font(size=11, color='0563C1', underline='single')
            if r % 2 == 0:
                lc.fill = _fill(CLR['alt'])
                vc.fill = _fill(CLR['alt'])
            r += 1

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 26
        ws.column_dimensions['D'].width = 28

    def _create_critical_items_sheet(self, writer, critical_items: List[Dict[str, Any]]):
        """Create Critical Items sheet with professional table"""
        df = pd.DataFrame(critical_items)
        
        # Ensure columns are in proper order
        column_order = ['type', 'repository', 'severity', 'description']
        df = df[column_order]
        
        # Rename for display
        df.columns = ['Alert Type', 'Repository', 'Severity', 'Description']
        
        # Sort by severity
        severity_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        df['_sort'] = df['Severity'].map(severity_order)
        df = df.sort_values('_sort').drop('_sort', axis=1)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Critical Items', index=False, startrow=0)
    
    def _create_exposed_secrets_sheet(self, writer, exposed_secrets: List[Dict[str, Any]]):
        """Create Exposed Secrets sheet"""
        df = pd.DataFrame(exposed_secrets)
        
        # Rename columns
        column_mapping = {
            'repository': 'Repository',
            'secret_type': 'Secret Type',
            'age_days': 'Age (Days)'
        }
        df = df.rename(columns=column_mapping)
        
        # Sort by age (oldest first)
        df = df.sort_values('Age (Days)', ascending=False)
        
        df.to_excel(writer, sheet_name='Exposed Secrets', index=False)
    
    def _create_summary_dashboard(self, writer, summary: Dict[str, Any]):
        """Create summary dashboard sheet"""
        critical_count = len(summary.get('critical_items', []))
        secrets_count = len(summary.get('exposed_secrets', []))
        
        # Count by type
        type_counts = {}
        severity_counts = {}
        repo_counts = {}
        
        for item in summary.get('critical_items', []):
            # Count by type
            item_type = item.get('type', 'Unknown')
            type_counts[item_type] = type_counts.get(item_type, 0) + 1
            
            # Count by severity
            severity = item.get('severity', 'unknown')
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
            
            # Count by repository
            repo = item.get('repository', 'Unknown')
            repo_counts[repo] = repo_counts.get(repo, 0) + 1
        
        # Create summary data
        summary_data = {
            'Metric': [
                'Report Date',
                'Total Critical Items',
                'Exposed Secrets',
                '',
                'By Alert Type:',
            ],
            'Value': [
                summary.get('date', datetime.now().strftime('%Y-%m-%d')),
                critical_count,
                secrets_count,
                '',
                ''
            ]
        }
        
        # Add type breakdowns
        for alert_type, count in sorted(type_counts.items()):
            summary_data['Metric'].append(f'  • {alert_type}')
            summary_data['Value'].append(count)
        
        summary_data['Metric'].append('')
        summary_data['Value'].append('')
        summary_data['Metric'].append('By Severity:')
        summary_data['Value'].append('')
        
        # Add severity breakdowns
        for severity, count in sorted(severity_counts.items()):
            summary_data['Metric'].append(f'  • {severity.title()}')
            summary_data['Value'].append(count)
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    def _apply_formatting(self, filename: Path, summary: Dict[str, Any]):
        """Apply professional formatting to all sheets"""
        wb = load_workbook(filename)

        # ── Pin Executive Summary as the first (leftmost) tab ─────────────────
        if 'Executive Summary' in wb.sheetnames:
            wb.move_sheet('Executive Summary',
                          offset=-wb.sheetnames.index('Executive Summary'))

        # ── Format data sheets (Executive Summary is already fully styled) ─────
        if 'Critical Items' in wb.sheetnames:
            self._format_critical_items_sheet(wb['Critical Items'])
            if summary.get('critical_items'):
                self._create_pivot_analysis(wb, summary['critical_items'])

        if 'Exposed Secrets' in wb.sheetnames:
            self._format_exposed_secrets_sheet(wb['Exposed Secrets'])

        wb.save(filename)
    
    def _format_critical_items_sheet(self, worksheet):
        """Apply professional formatting to Critical Items sheet"""
        # Header row styling
        header_fill = PatternFill(start_color=self.colors['header_bg'], 
                                  end_color=self.colors['header_bg'], 
                                  fill_type='solid')
        header_font = Font(bold=True, color=self.colors['header_text'], size=12)
        
        # Apply to header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code severity column
        severity_col = 3  # Column C (Severity)
        
        for row in range(2, worksheet.max_row + 1):
            severity_cell = worksheet.cell(row=row, column=severity_col)
            severity = str(severity_cell.value).lower() if severity_cell.value else ''
            
            if 'critical' in severity:
                severity_cell.fill = PatternFill(start_color=self.colors['critical'], 
                                                end_color=self.colors['critical'], 
                                                fill_type='solid')
                severity_cell.font = Font(bold=True, color='FFFFFF')
            elif 'high' in severity:
                severity_cell.fill = PatternFill(start_color=self.colors['high'], 
                                                end_color=self.colors['high'], 
                                                fill_type='solid')
                severity_cell.font = Font(bold=True, color='FFFFFF')
            elif 'medium' in severity:
                severity_cell.fill = PatternFill(start_color=self.colors['medium'], 
                                                end_color=self.colors['medium'], 
                                                fill_type='solid')
                severity_cell.font = Font(bold=True, color='000000')
            elif 'low' in severity:
                severity_cell.fill = PatternFill(start_color=self.colors['low'], 
                                                end_color=self.colors['low'], 
                                                fill_type='solid')
                severity_cell.font = Font(bold=True, color='000000')
            
            # Alternate row colors for readability
            if row % 2 == 0:
                for col in range(1, worksheet.max_column + 1):
                    if col != severity_col:  # Don't overwrite severity colors
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color=self.colors['alt_row'], 
                                              end_color=self.colors['alt_row'], 
                                              fill_type='solid')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 80)  # Cap at 80
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
    
    def _format_exposed_secrets_sheet(self, worksheet):
        """Apply formatting to Exposed Secrets sheet"""
        # Header row
        header_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Highlight old secrets (>30 days)
        age_col = 3  # Age (Days) column
        
        for row in range(2, worksheet.max_row + 1):
            age_cell = worksheet.cell(row=row, column=age_col)
            
            try:
                age = int(age_cell.value) if age_cell.value else 0
                if age > 30:
                    age_cell.fill = PatternFill(start_color='FF6B6B', 
                                               end_color='FF6B6B', 
                                               fill_type='solid')
                    age_cell.font = Font(bold=True, color='FFFFFF')
            except:
                pass
        
        # Auto-adjust columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_summary_sheet(self, worksheet):
        """Apply formatting to Summary sheet"""
        # Title formatting
        for row in range(1, worksheet.max_row + 1):
            cell_a = worksheet.cell(row=row, column=1)
            cell_b = worksheet.cell(row=row, column=2)
            
            # Bold metrics that end with ":"
            if cell_a.value and str(cell_a.value).endswith(':'):
                cell_a.font = Font(bold=True, size=12, color='1E2761')
            
            # Highlight totals
            if 'Total' in str(cell_a.value):
                cell_a.font = Font(bold=True, size=12)
                cell_b.font = Font(bold=True, size=12)
                cell_a.fill = PatternFill(start_color='E8F4F8', 
                                         end_color='E8F4F8', 
                                         fill_type='solid')
                cell_b.fill = PatternFill(start_color='E8F4F8', 
                                         end_color='E8F4F8', 
                                         fill_type='solid')
        
        # Column widths
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 15
    
    def _create_pivot_analysis(self, workbook, critical_items: List[Dict[str, Any]]):
        """Create pivot table analysis sheet"""
        # Create new sheet for pivot analysis
        if 'Pivot Analysis' in workbook.sheetnames:
            del workbook['Pivot Analysis']
        
        ws_pivot = workbook.create_sheet('Pivot Analysis')
        
        # Title
        ws_pivot['A1'] = 'Critical Items Analysis'
        ws_pivot['A1'].font = Font(bold=True, size=16, color='1E2761')
        ws_pivot.merge_cells('A1:D1')
        
        # Group by repository
        repo_counts = {}
        type_counts = {}
        severity_counts = {}
        
        for item in critical_items:
            repo = item.get('repository', 'Unknown')
            alert_type = item.get('type', 'Unknown')
            severity = item.get('severity', 'unknown')
            
            repo_counts[repo] = repo_counts.get(repo, 0) + 1
            type_counts[alert_type] = type_counts.get(alert_type, 0) + 1
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        # Section 1: By Repository
        current_row = 3
        ws_pivot[f'A{current_row}'] = 'Critical Items by Repository'
        ws_pivot[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1
        
        ws_pivot[f'A{current_row}'] = 'Repository'
        ws_pivot[f'B{current_row}'] = 'Count'
        ws_pivot[f'A{current_row}'].font = Font(bold=True)
        ws_pivot[f'B{current_row}'].font = Font(bold=True)
        ws_pivot[f'A{current_row}'].fill = PatternFill(start_color='1E2761', 
                                                        end_color='1E2761', 
                                                        fill_type='solid')
        ws_pivot[f'B{current_row}'].fill = PatternFill(start_color='1E2761', 
                                                        end_color='1E2761', 
                                                        fill_type='solid')
        ws_pivot[f'A{current_row}'].font = Font(bold=True, color='FFFFFF')
        ws_pivot[f'B{current_row}'].font = Font(bold=True, color='FFFFFF')
        current_row += 1
        
        for repo, count in sorted(repo_counts.items(), key=lambda x: x[1], reverse=True):
            ws_pivot[f'A{current_row}'] = repo
            ws_pivot[f'B{current_row}'] = count
            current_row += 1
        
        # Section 2: By Alert Type
        current_row += 2
        ws_pivot[f'A{current_row}'] = 'Critical Items by Type'
        ws_pivot[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1
        
        ws_pivot[f'A{current_row}'] = 'Alert Type'
        ws_pivot[f'B{current_row}'] = 'Count'
        ws_pivot[f'A{current_row}'].font = Font(bold=True)
        ws_pivot[f'B{current_row}'].font = Font(bold=True)
        ws_pivot[f'A{current_row}'].fill = PatternFill(start_color='1E2761', 
                                                        end_color='1E2761', 
                                                        fill_type='solid')
        ws_pivot[f'B{current_row}'].fill = PatternFill(start_color='1E2761', 
                                                        end_color='1E2761', 
                                                        fill_type='solid')
        ws_pivot[f'A{current_row}'].font = Font(bold=True, color='FFFFFF')
        ws_pivot[f'B{current_row}'].font = Font(bold=True, color='FFFFFF')
        current_row += 1
        
        for alert_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            ws_pivot[f'A{current_row}'] = alert_type
            ws_pivot[f'B{current_row}'] = count
            current_row += 1
        
        # Section 3: By Severity
        current_row += 2
        ws_pivot[f'A{current_row}'] = 'Critical Items by Severity'
        ws_pivot[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1
        
        ws_pivot[f'A{current_row}'] = 'Severity'
        ws_pivot[f'B{current_row}'] = 'Count'
        ws_pivot[f'A{current_row}'].font = Font(bold=True)
        ws_pivot[f'B{current_row}'].font = Font(bold=True)
        ws_pivot[f'A{current_row}'].fill = PatternFill(start_color='1E2761', 
                                                        end_color='1E2761', 
                                                        fill_type='solid')
        ws_pivot[f'B{current_row}'].fill = PatternFill(start_color='1E2761', 
                                                        end_color='1E2761', 
                                                        fill_type='solid')
        ws_pivot[f'A{current_row}'].font = Font(bold=True, color='FFFFFF')
        ws_pivot[f'B{current_row}'].font = Font(bold=True, color='FFFFFF')
        current_row += 1
        
        severity_order = ['critical', 'high', 'medium', 'low']
        for severity in severity_order:
            if severity in severity_counts:
                ws_pivot[f'A{current_row}'] = severity.title()
                ws_pivot[f'B{current_row}'] = severity_counts[severity]
                
                # Color code
                if severity == 'critical':
                    color = 'DC143C'
                elif severity == 'high':
                    color = 'FF8C00'
                elif severity == 'medium':
                    color = 'FFD700'
                else:
                    color = '90EE90'
                
                ws_pivot[f'A{current_row}'].fill = PatternFill(start_color=color, 
                                                                end_color=color, 
                                                                fill_type='solid')
                ws_pivot[f'B{current_row}'].fill = PatternFill(start_color=color, 
                                                                end_color=color, 
                                                                fill_type='solid')
                ws_pivot[f'A{current_row}'].font = Font(bold=True, color='FFFFFF' if severity in ['critical', 'high'] else '000000')
                ws_pivot[f'B{current_row}'].font = Font(bold=True, color='FFFFFF' if severity in ['critical', 'high'] else '000000')
                
                current_row += 1
        
        # Adjust column widths
        ws_pivot.column_dimensions['A'].width = 35
        ws_pivot.column_dimensions['B'].width = 15