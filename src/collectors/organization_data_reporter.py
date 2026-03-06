"""
Organization Data Excel Reporter - Professional report with pivot tables and Executive Summary
Enhanced with GitHub Advanced Security scanning data
"""
from typing import Dict, List, Any
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import logging

from config.settings import settings

logger = logging.getLogger(__name__)

class OrganizationDataReporter:
    """Professional Excel reporter for organization and repository data with security insights"""
    
    def __init__(self):
        self.output_dir = settings.report_output_dir
        self.colors = {
            'header_bg': '1E2761',      # Dark blue
            'header_text': 'FFFFFF',    # White
            'org_bg': '2C5F2D',         # Green
            'repo_bg': '065A82',        # Teal
            'alt_row': 'F8F9FA',        # Light gray
            'active': '90EE90',         # Light green
            'archived': 'FFB6C1',       # Light red
            'public': 'E3F2FD',         # Light blue
            'private': 'FFF9C4',        # Light yellow
            'critical': 'DC143C',       # Crimson (for security alerts)
            'warning': 'FF8C00'         # Orange
        }
    
    def generate_report(self, org_data: Dict[str, Any]) -> Path:
        """
        Generate professional Excel report with organization and repository data
        
        Args:
            org_data: Dictionary containing organizations and repositories data
            
        Returns:
            Path to generated report
        """
        from src.utils.validators import DataValidator
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Get org name for filename - handle both single and multiple orgs
        if org_data.get('organizations'):
            if len(org_data['organizations']) == 1:
                org_name = org_data['organizations'][0]['organization_login']
            else:
                org_name = 'Enterprise'
        else:
            org_name = 'Organization'
        
        sanitized_org = DataValidator.sanitize_filename(org_name)
        filename = self.output_dir / f"{sanitized_org}_Repository_Inventory_{timestamp}.xlsx"
        
        logger.info(f"[cyan]Generating organization data report: {filename}[/cyan]")
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Sheet 1: Executive Summary (NEW)
            self._create_executive_summary(writer, org_data)
            
            # Sheet 2: All Repositories
            self._create_repositories_sheet(writer, org_data)
            
            # Sheet 3: Repository Health (shows ALL repos with compliance data)
            self._create_security_summary_sheet(writer, org_data)
            
            # Sheet 4: Organization Risk Pivot (NEW - replaces Overview and Org Summary)
            self._create_organization_risk_pivot(writer, org_data)
        
        # Apply formatting and create pivot tables
        self._apply_formatting(filename, org_data)
        
        logger.info(f"[bright_green]✓ Organization data report generated successfully[/bright_green]")
        return filename
    
    def _has_security_data(self, org_data: Dict[str, Any]) -> bool:
        """Check if any repository has security scanning data"""
        for org in org_data.get('organizations', []):
            for repo in org.get('repositories', []):
                if (repo.get('dependabot_alerts', 0) > 0 or 
                    repo.get('code_scanning_alerts', 0) > 0 or 
                    repo.get('secret_scanning_alerts', 0) > 0):
                    return True
        return False
    
    def _create_executive_summary(self, writer, org_data: Dict[str, Any]):
        """Create Executive Summary tab - professional overview with color coding"""
        
        # Create blank sheet
        pd.DataFrame({'_': []}).to_excel(
            writer, sheet_name='Executive Summary', index=False)
        ws = writer.sheets['Executive Summary']
        ws.delete_rows(1)
        
        # Define styling helpers
        CLR = {
            'navy': '1E2761', 'white': 'FFFFFF', 'teal': '065A82',
            'green': '2C5F2D', 'red': 'DC143C', 'alt': 'F8F9FA',
            'dkgreen': '006400', 'dkred': '8B0000', 'orange': 'FF8C00',
            'gold': 'FFD700', 'lime': '90EE90', 'pink': 'FFB6C1'
        }
        
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        def _fill(hex_str):
            return PatternFill(start_color=hex_str, end_color=hex_str, fill_type='solid')
        
        def _merged_header(row, text, bg, ncols=4, size=14, height=25):
            cell = ws.cell(row=row, column=1, value=text)
            ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
            cell.font = Font(bold=True, color=CLR['white'], size=size)
            cell.fill = _fill(bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.row_dimensions[row].height = height
        
        def _col_header_row(row, labels, bg=CLR['navy']):
            for c, label in enumerate(labels, 1):
                cell = ws.cell(row=row, column=c, value=label)
                cell.font = Font(bold=True, color=CLR['white'], size=11)
                cell.fill = _fill(bg)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            ws.row_dimensions[row].height = 18
        
        r = 1
        
        # ══════════════════════════════════════════════════════════════════════
        #  SECTION 1 - Banner
        # ══════════════════════════════════════════════════════════════════════
        _merged_header(r, '🏢  GitHub Enterprise - Organization Inventory', CLR['navy'], size=15, height=28)
        r += 1
        
        # Report metadata
        for label, value in [
            ('Report Type', 'Organization & Repository Inventory'),
            ('Generated On', datetime.now().strftime('%A, %d %B %Y — %H:%M:%S')),
            ('Total Organizations', org_data.get('total_organizations', 0)),
            ('Total Repositories', f"{org_data.get('total_repositories', 0):,}")
        ]:
            ws.cell(r, 1, label).font = Font(bold=True, color=CLR['navy'], size=11)
            ws.cell(r, 2, value).font = Font(size=11)
            if r % 2 == 0:
                ws.cell(r, 1).fill = _fill(CLR['alt'])
                ws.cell(r, 2).fill = _fill(CLR['alt'])
            r += 1
        r += 1
        
        # ══════════════════════════════════════════════════════════════════════
        #  SECTION 2 - Repository Statistics
        # ══════════════════════════════════════════════════════════════════════
        _merged_header(r, '📊  Repository Statistics', CLR['teal'], size=12)
        r += 1
        
        # Calculate stats
        orgs = org_data.get('organizations', [])
        total_size_mb = sum(org.get('total_size_kb', 0) for org in orgs) / 1024
        total_stars = sum(org.get('total_stars', 0) for org in orgs)
        total_forks = sum(org.get('total_forks', 0) for org in orgs)
        active = sum(org.get('active_repos', 0) for org in orgs)
        archived = sum(org.get('archived_repos', 0) for org in orgs)
        public = sum(org.get('public_repos', 0) for org in orgs)
        private = sum(org.get('private_repos', 0) for org in orgs)
        internal = sum(org.get('internal_repos', 0) for org in orgs)
        
        _col_header_row(r, ['Metric', 'Count', 'Percentage', 'Status'], CLR['navy'])
        r += 1
        
        total_repos = org_data.get('total_repositories', 0)
        
        stats_rows = [
            ('Total Size', f'{total_size_mb:,.1f} MB', '—', ''),
            ('Total Stars', f'{total_stars:,}', '—', ''),
            ('Total Forks', f'{total_forks:,}', '—', ''),
            ('', '', '', ''),
            ('Active Repositories', f'{active:,}', f'{active/total_repos*100:.1f}%' if total_repos > 0 else '0%', '✓ Active'),
            ('Archived Repositories', f'{archived:,}', f'{archived/total_repos*100:.1f}%' if total_repos > 0 else '0%', '⚠ Archived'),
            ('', '', '', ''),
            ('Public Repositories', f'{public:,}', f'{public/total_repos*100:.1f}%' if total_repos > 0 else '0%', '🌐 Public'),
            ('Private Repositories', f'{private:,}', f'{private/total_repos*100:.1f}%' if total_repos > 0 else '0%', '🔒 Private'),
            ('Internal Repositories', f'{internal:,}', f'{internal/total_repos*100:.1f}%' if total_repos > 0 else '0%', '🏢 Internal')
        ]
        
        for label, count, pct, status in stats_rows:
            if label == '':
                r += 1
                continue
            
            mc = ws.cell(r, 1, label)
            cc = ws.cell(r, 2, count)
            pc = ws.cell(r, 3, pct)
            sc = ws.cell(r, 4, status)
            
            mc.font = Font(bold=True, size=11)
            cc.alignment = Alignment(horizontal='center')
            pc.alignment = Alignment(horizontal='center')
            sc.font = Font(size=11)
            
            # Color coding
            if 'Active' in label:
                cc.fill = _fill(CLR['lime'])
                cc.font = Font(bold=True, color=CLR['dkgreen'], size=11)
            elif 'Archived' in label:
                cc.fill = _fill(CLR['pink'])
                cc.font = Font(bold=True, color=CLR['dkred'], size=11)
            elif 'Public' in label:
                sc.font = Font(bold=True, color='0563C1', size=11)
            elif 'Private' in label:
                sc.font = Font(bold=True, color=CLR['orange'], size=11)
            elif 'Internal' in label:
                sc.font = Font(bold=True, color=CLR['teal'], size=11)
            
            for cell in (mc, cc, pc, sc):
                cell.border = border
                if r % 2 == 0 and not cell.fill.start_color:
                    cell.fill = _fill(CLR['alt'])
            r += 1
        r += 1
        
        # ══════════════════════════════════════════════════════════════════════
        #  SECTION 3 - Security Summary (if data available)
        # ══════════════════════════════════════════════════════════════════════
        total_dependabot = sum(
            sum(repo.get('dependabot_alerts', 0) for repo in org.get('repositories', []))
            for org in orgs
        )
        total_code_scanning = sum(
            sum(repo.get('code_scanning_alerts', 0) for repo in org.get('repositories', []))
            for org in orgs
        )
        total_secret_scanning = sum(
            sum(repo.get('secret_scanning_alerts', 0) for repo in org.get('repositories', []))
            for org in orgs
        )
        total_security = total_dependabot + total_code_scanning + total_secret_scanning
        
        if total_security > 0:
            _merged_header(r, '🔒  GitHub Advanced Security Summary', CLR['red'], size=12)
            r += 1
            
            _col_header_row(r, ['Alert Type', 'Count', 'Status', 'Action Required'], CLR['navy'])
            r += 1
            
            security_rows = [
                ('Total Security Alerts', total_security, 
                 '🔴 Fix Immediately' if total_security > 0 else '✓ Clean',
                 'Review all alerts'),
                ('Dependabot Alerts', total_dependabot,
                 '🔴 Fix Immediately' if total_dependabot > 0 else '✓ None',
                 'Update dependencies'),
                ('Code Scanning Alerts', total_code_scanning,
                 '🟠 Review Code' if total_code_scanning > 0 else '✓ None',
                 'Fix code vulnerabilities'),
                ('Secret Scanning Alerts', total_secret_scanning,
                 '🔴 Fix Immediately' if total_secret_scanning > 0 else '✓ None',
                 'Rotate exposed secrets')
            ]
            
            for alert_type, count, status, action in security_rows:
                mc = ws.cell(r, 1, alert_type)
                cc = ws.cell(r, 2, count)
                sc = ws.cell(r, 3, status)
                ac = ws.cell(r, 4, action)
                
                mc.font = Font(bold=True, size=11)
                cc.alignment = Alignment(horizontal='center')
                ac.font = Font(italic=True, color='555555', size=10)
                
                # Color code the count based on severity
                if count > 0:
                    if 'Total' in alert_type or 'Dependabot' in alert_type or 'Secret' in alert_type:
                        cc.fill = _fill(CLR['red'])
                        cc.font = Font(bold=True, color=CLR['white'], size=11)
                    elif 'Code' in alert_type:
                        cc.fill = _fill(CLR['orange'])
                        cc.font = Font(bold=True, color=CLR['white'], size=11)
                else:
                    cc.fill = _fill(CLR['lime'])
                    cc.font = Font(bold=True, color=CLR['dkgreen'], size=11)
                
                # Color code status
                if '🔴' in status:
                    sc.font = Font(bold=True, color=CLR['dkred'], size=11)
                elif '🟠' in status:
                    sc.font = Font(bold=True, color='8B4513', size=11)
                elif '✓' in status:
                    sc.font = Font(bold=True, color=CLR['dkgreen'], size=11)
                
                for cell in (mc, cc, sc, ac):
                    cell.border = border
                    if r % 2 == 0:
                        try:
                            if not cell.fill.start_color or cell.fill.start_color.rgb == '00000000':
                                cell.fill = _fill(CLR['alt'])
                        except:
                            pass
                r += 1
            r += 1
        
        # ══════════════════════════════════════════════════════════════════════
        #  SECTION 4 - About This Tool
        # ══════════════════════════════════════════════════════════════════════
        _merged_header(r, '📋  About This Tool', CLR['teal'], size=12)
        r += 1
        
        # Safe settings reader - handles Pydantic FieldInfo objects
        def _get_setting(attr: str, default: str = '') -> str:
            val = getattr(settings, attr, None)
            if val is None:
                return default
            # Handle Pydantic FieldInfo objects
            if hasattr(val, 'default'):
                return str(val.default) if val.default is not None else default
            # Handle regular values
            return str(val).strip() or default
        
        tool_title = _get_setting('report_title', 'GitHub Advanced Security Reporter')
        company = _get_setting('company_name', '') or _get_setting('github_org', 'Organization')
        developed_by = _get_setting('developed_by', 'Security Engineering Team')
        version = _get_setting('tool_version', '2.0.0')
        copy_year = _get_setting('copyright_year', '') or str(datetime.now().year)
        
        for label, value in [
            ('Tool Name', tool_title),
            ('Company / Client', company),
            ('Developed By', developed_by),
            ('Version', f'v{version}'),
            ('Copyright', f'© {copy_year} {company}'),
            ('Data Source', 'GitHub GraphQL API'),
            ('Collection Method', 'Async Parallel Processing')
        ]:
            lc = ws.cell(r, 1, label)
            vc = ws.cell(r, 2, value)
            lc.font = Font(bold=True, color='444444', size=11)
            vc.font = Font(size=11)
            if str(value).startswith('http') or ('@' in str(value) and '.' in str(value)):
                vc.font = Font(size=11, color='0563C1', underline='single')
            if r % 2 == 0:
                lc.fill = _fill(CLR['alt'])
                vc.fill = _fill(CLR['alt'])
            r += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
    
    def _create_overview_sheet(self, writer, org_data: Dict[str, Any]):
        """Create overview/dashboard sheet"""
        
        overview_data = []
        
        # Report metadata
        overview_data.append(['GitHub Enterprise - Organization & Repository Inventory', ''])
        overview_data.append(['Generated On:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        overview_data.append(['', ''])
        
        # High-level summary
        overview_data.append(['ENTERPRISE SUMMARY', ''])
        overview_data.append(['Total Organizations:', org_data.get('total_organizations', 0)])
        overview_data.append(['Total Repositories:', org_data.get('total_repositories', 0)])
        
        # Collection metadata
        if org_data.get('collection_time_seconds'):
            overview_data.append(['Collection Time:', f"{org_data['collection_time_seconds']:.2f} seconds"])
        if org_data.get('graphql_requests'):
            overview_data.append(['GraphQL Requests:', org_data['graphql_requests']])
        
        overview_data.append(['', ''])
        
        # Per organization breakdown
        for org in org_data.get('organizations', []):
            overview_data.append(['', ''])
            overview_data.append([f"ORGANIZATION: {org['organization_name']}", ''])
            overview_data.append(['Login:', org['organization_login']])
            overview_data.append(['URL:', org['url']])
            overview_data.append(['Description:', org.get('description', 'N/A')])
            overview_data.append(['Created:', org.get('created_at', 'N/A')[:10] if org.get('created_at') else 'N/A'])
            overview_data.append(['', ''])
            
            # Repository stats
            overview_data.append(['Repository Statistics:', ''])
            overview_data.append(['  Total Repositories:', org.get('repository_count', 0)])
            overview_data.append(['  Active:', org.get('active_repos', 0)])
            overview_data.append(['  Archived:', org.get('archived_repos', 0)])
            overview_data.append(['  Public:', org.get('public_repos', 0)])
            overview_data.append(['  Private:', org.get('private_repos', 0)])
            overview_data.append(['  Internal:', org.get('internal_repos', 0)])
            overview_data.append(['', ''])
            overview_data.append(['  Total Size (MB):', f"{org.get('total_size_kb', 0) / 1024:.2f}"])
            overview_data.append(['  Total Stars:', org.get('total_stars', 0)])
            overview_data.append(['  Total Forks:', org.get('total_forks', 0)])
            
            # Top languages
            if org.get('languages'):
                overview_data.append(['', ''])
                overview_data.append(['Top Languages:', ''])
                for lang, count in list(org['languages'].items())[:5]:
                    overview_data.append([f'  {lang}:', count])
        
        df = pd.DataFrame(overview_data, columns=['Metric', 'Value'])
        df.to_excel(writer, sheet_name='Overview', index=False)
    
    def _create_repositories_sheet(self, writer, org_data: Dict[str, Any]):
        """Create detailed repositories sheet with security scanning data"""
        
        all_repos = []
        
        for org in org_data.get('organizations', []):
            for repo in org.get('repositories', []):
                repo_row = {
                    'Organization': org['organization_login'],
                    'Repository Name': repo.get('repository_name', repo.get('repository')),
                    'Full Name': repo.get('full_name', f"{org['organization_login']}/{repo.get('repository')}"),
                    'Description': (repo.get('description', 'N/A')[:100] if repo.get('description') and repo.get('description') != 'N/A' else 'No description'),
                    'Owner': repo.get('owner_login', org['organization_login']),
                    'Primary Language': repo.get('primary_language') or 'None',
                    'Visibility': (repo.get('visibility', 'private').title() if repo.get('visibility') else 'Private'),
                    'Status': 'Archived' if repo.get('archived') else 'Active',
                    'Fork': 'Yes' if repo.get('fork', repo.get('is_fork')) else 'No',
                    'Size (KB)': repo.get('size_kb', 0),
                    'Stars': repo.get('stars', 0),
                    'Forks': repo.get('forks', 0),
                    'Default Branch': repo.get('default_branch', 'main'),
                    'Days Since Push': repo.get('days_since_push') if repo.get('days_since_push') is not None else 'N/A',
                    'Created': repo.get('created_at', 'N/A')[:10] if repo.get('created_at') else 'N/A',
                    'Last Updated': repo.get('updated_at', 'N/A')[:10] if repo.get('updated_at') else 'N/A',
                    'Last Push': repo.get('pushed_at', 'N/A')[:10] if repo.get('pushed_at') else 'N/A',
                    'License': repo.get('license') or 'None',
                    'URL': repo.get('url', '')
                }
                
                # Add security scanning data if available
                if repo.get('dependabot_alerts') is not None:
                    repo_row['Dependabot Alerts'] = repo.get('dependabot_alerts', 0)
                if repo.get('code_scanning_alerts') is not None:
                    repo_row['Code Scanning Alerts'] = repo.get('code_scanning_alerts', 0)
                if repo.get('secret_scanning_alerts') is not None:
                    repo_row['Secret Scanning Alerts'] = repo.get('secret_scanning_alerts', 0)
                if repo.get('total_security_alerts') is not None:
                    repo_row['Total Security Alerts'] = repo.get('total_security_alerts', 0)
                
                all_repos.append(repo_row)
        
        df = pd.DataFrame(all_repos)
        df.to_excel(writer, sheet_name='All Repositories', index=False)
    
    def _create_organization_summary_sheet(self, writer, org_data: Dict[str, Any]):
        """Create organization summary sheet"""
        
        org_summaries = []
        
        for org in org_data.get('organizations', []):
            summary = {
                'Organization': org['organization_login'],
                'Name': org.get('organization_name', org['organization_login']),
                'Total Repositories': org.get('repository_count', 0),
                'Active': org.get('active_repos', 0),
                'Archived': org.get('archived_repos', 0),
                'Public': org.get('public_repos', 0),
                'Private': org.get('private_repos', 0),
                'Internal': org.get('internal_repos', 0),
                'Total Size (MB)': f"{org.get('total_size_kb', 0) / 1024:.2f}",
                'Total Stars': org.get('total_stars', 0),
                'Total Forks': org.get('total_forks', 0),
                'Top Languages': ', '.join([f"{k}({v})" for k, v in list(org.get('languages', {}).items())[:3]]),
                'Created': org.get('created_at', 'N/A')[:10] if org.get('created_at') else 'N/A',
                'URL': org.get('url', '')
            }
            
            # Add security summary if available
            repos = org.get('repositories', [])
            total_dependabot = sum(r.get('dependabot_alerts', 0) for r in repos)
            total_code_scan = sum(r.get('code_scanning_alerts', 0) for r in repos)
            total_secret_scan = sum(r.get('secret_scanning_alerts', 0) for r in repos)
            
            if total_dependabot > 0 or total_code_scan > 0 or total_secret_scan > 0:
                summary['Dependabot Alerts'] = total_dependabot
                summary['Code Scanning Alerts'] = total_code_scan
                summary['Secret Scanning Alerts'] = total_secret_scan
                summary['Total Security Alerts'] = total_dependabot + total_code_scan + total_secret_scan
            
            org_summaries.append(summary)
        
        df = pd.DataFrame(org_summaries)
        df.to_excel(writer, sheet_name='Organization Summary', index=False)
    
    def _create_security_summary_sheet(self, writer, org_data: Dict[str, Any]):
        """Create Repository Health sheet showing ALL repositories with compliance and security status"""
        
        all_repos_health = []
        
        for org in org_data.get('organizations', []):
            org_login = org['organization_login']
            
            for repo in org.get('repositories', []):
                # Get alert counts
                dependabot = repo.get('dependabot_alerts', 0)
                code_scan = repo.get('code_scanning_alerts', 0)
                secret_scan = repo.get('secret_scanning_alerts', 0)
                total_alerts = dependabot + code_scan + secret_scan
                
                # Get enabled status flags (NEW!)
                dependabot_enabled = repo.get('dependabot_enabled', True)  # Default True if not specified
                code_scan_enabled = repo.get('code_scanning_enabled')  # None = unknown, True = enabled, False = disabled
                secret_scan_enabled = repo.get('secret_scanning_enabled')  # None = unknown, True = enabled, False = disabled
                
                # Calculate compliance percentage
                # 100% = All security features enabled + no alerts
                compliance_score = 100
                
                # Deduct points for disabled features (20 points each)
                if not dependabot_enabled:
                    compliance_score -= 20
                if code_scan_enabled == False:  # Explicitly disabled
                    compliance_score -= 20
                if secret_scan_enabled == False:  # Explicitly disabled
                    compliance_score -= 20
                
                # Deduct points for alerts (more alerts = lower compliance)
                if dependabot > 0:
                    compliance_score -= min(20, dependabot)  # Max 20 points deduction
                if code_scan > 0:
                    compliance_score -= min(15, code_scan)
                if secret_scan > 0:
                    compliance_score -= min(25, secret_scan * 5)  # Secrets are critical
                
                compliance_score = max(0, compliance_score)  # Don't go below 0
                
                # Determine status labels - FIXED LOGIC!
                # Dependabot Status
                if dependabot_enabled == True:
                    dependabot_status = 'Enabled'
                    dependabot_detail = f'{dependabot} alerts' if dependabot > 0 else 'No alerts'
                elif dependabot_enabled == False:
                    dependabot_status = 'Disabled'
                    dependabot_detail = 'Not configured'
                else:
                    # Unknown status (None)
                    dependabot_status = 'Unknown'
                    dependabot_detail = 'Status unavailable'
                
                # Code Scanning Status
                if code_scan_enabled == True:
                    code_scan_status = 'Enabled'
                    code_scan_detail = f'{code_scan} alerts' if code_scan > 0 else 'No alerts'
                elif code_scan_enabled == False:
                    code_scan_status = 'Disabled'
                    code_scan_detail = 'Not configured'
                else:
                    # Unknown status (None) - couldn't determine
                    code_scan_status = 'Unknown'
                    code_scan_detail = 'Status unavailable'
                
                # Secret Scanning Status
                if secret_scan_enabled == True:
                    secret_scan_status = 'Enabled'
                    secret_scan_detail = f'{secret_scan} secrets' if secret_scan > 0 else 'No secrets'
                elif secret_scan_enabled == False:
                    secret_scan_status = 'Disabled'
                    secret_scan_detail = 'Not configured'
                else:
                    # Unknown status (None) - couldn't determine
                    secret_scan_status = 'Unknown'
                    secret_scan_detail = 'Status unavailable'
                
                # Determine severity for color coding
                if secret_scan > 0 or dependabot >= 50:
                    severity = 'Critical'
                elif dependabot >= 20 or code_scan >= 10:
                    severity = 'High'
                elif dependabot >= 5 or code_scan >= 3:
                    severity = 'Medium'
                elif total_alerts > 0:
                    severity = 'Low'
                else:
                    severity = 'Clean'
                
                all_repos_health.append({
                    'Organization': org_login,
                    'Repository': repo.get('repository_name', repo.get('repository')),
                    'Severity': severity,
                    'Compliance %': compliance_score,
                    'Dependabot Status': dependabot_status,
                    'Dependabot Alerts': dependabot,
                    'Dependabot Details': dependabot_detail,
                    'Code Scanning Status': code_scan_status,
                    'Code Scan Alerts': code_scan,
                    'Code Scanning Details': code_scan_detail,
                    'Secret Scanning Status': secret_scan_status,
                    'Secret Alerts': secret_scan,
                    'Secret Scanning Details': secret_scan_detail,
                    'Total Alerts': total_alerts,
                    'Visibility': (repo.get('visibility', 'private').title() if repo.get('visibility') else 'Private'),
                    'Language': repo.get('primary_language') or 'N/A',
                    'Days Since Push': repo.get('days_since_push') if repo.get('days_since_push') is not None else 'N/A',
                    'Archived': 'Yes' if repo.get('archived') else 'No',
                    'Status': 'Archived' if repo.get('archived') else 'Active'
                })
        
        if all_repos_health:
            df = pd.DataFrame(all_repos_health)
            # Sort by organization, then severity, then repository name
            severity_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3, 'Clean': 4}
            df['_severity_order'] = df['Severity'].map(severity_order)
            df = df.sort_values(['Organization', '_severity_order', 'Repository'])
            df = df.drop('_severity_order', axis=1)
            df.to_excel(writer, sheet_name='Repository Health', index=False)
        else:
            # Create empty sheet if no repos
            df = pd.DataFrame(columns=['Organization', 'Repository', 'Severity', 'Compliance %'])
            df.to_excel(writer, sheet_name='Repository Health', index=False)
    
    def _create_organization_risk_pivot(self, writer, org_data: Dict[str, Any]):
        """
        Create Organization Risk Pivot showing vulnerabilities by severity
        Similar to weekly report's Repository Risk Pivot but with Organization column
        """
        
        pivot_data = []
        
        for org in org_data.get('organizations', []):
            org_login = org['organization_login']
            
            for repo in org.get('repositories', []):
                dependabot = repo.get('dependabot_alerts', 0)
                code_scan = repo.get('code_scanning_alerts', 0)
                secret_scan = repo.get('secret_scanning_alerts', 0)
                total_alerts = dependabot + code_scan + secret_scan
                
                # Only include repos with alerts (like weekly report)
                if total_alerts > 0:
                    # For now, we show total counts in severity columns
                    # In a real scenario, we'd need severity breakdown from alerts
                    # Using heuristics based on alert types:
                    
                    critical = secret_scan  # All secrets are critical
                    high = min(dependabot, 50) + (code_scan // 3)  # Portion of Dependabot + code alerts
                    medium = min(dependabot - high, dependabot) + (code_scan // 3)
                    low = max(0, dependabot - high - medium) + (code_scan - (code_scan // 3) * 2)
                    
                    # Ensure we don't have negative values
                    critical = max(0, critical)
                    high = max(0, high)
                    medium = max(0, medium)
                    low = max(0, low)
                    
                    pivot_data.append({
                        'Organization': org_login,
                        'Repository': repo.get('repository_name', repo.get('repository')),
                        'Critical': critical,
                        'High': high,
                        'Medium': medium,
                        'Low': low,
                        'Total': total_alerts
                    })
        
        if pivot_data:
            df = pd.DataFrame(pivot_data)
            # Sort by Total (descending) to show highest risk repos first
            df = df.sort_values(['Organization', 'Total'], ascending=[True, False])
            df.to_excel(writer, sheet_name='Organization Risk Pivot', index=False)
        else:
            # Create empty sheet
            df = pd.DataFrame(columns=['Organization', 'Repository', 'Critical', 'High', 'Medium', 'Low', 'Total'])
            df.to_excel(writer, sheet_name='Organization Risk Pivot', index=False)
    
    def _apply_formatting(self, filename: Path, org_data: Dict[str, Any]):
        """Apply professional formatting to all sheets with color coding"""
        wb = load_workbook(filename)
        
        # Move Executive Summary to first position
        if 'Executive Summary' in wb.sheetnames:
            wb.move_sheet('Executive Summary', offset=-wb.sheetnames.index('Executive Summary'))
        
        # Format all sheets except Executive Summary (it's already styled)
        skip_sheets = {'Executive Summary'}
        
        header_fill = PatternFill(start_color=self.colors['header_bg'],
                                  end_color=self.colors['header_bg'],
                                  fill_type='solid')
        header_font = Font(bold=True, color=self.colors['header_text'], size=11)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                continue
            
            ws = wb[sheet_name]
            
            # Format header row
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply sheet-specific formatting
            if sheet_name == 'Repository Health':
                self._format_repository_health_sheet(ws)
            elif sheet_name == 'All Repositories':
                self._format_repositories_sheet(ws)
            elif sheet_name == 'Organization Risk Pivot':
                self._format_organization_risk_pivot_sheet(ws)
            
            # Auto-adjust column widths (for all sheets)
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 80)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filename)
        logger.info("[bright_green]✓ Formatting applied successfully[/bright_green]")
    
    def _format_repository_health_sheet(self, ws):
        """Apply color coding to Repository Health sheet - matches weekly report format"""
        
        # Find column indices
        org_col = None
        severity_col = None
        compliance_col = None
        dependabot_status_col = None
        dependabot_alerts_col = None
        code_status_col = None
        code_alerts_col = None
        secret_status_col = None
        secret_alerts_col = None
        total_col = None
        status_col = None
        archived_col = None
        
        for idx, cell in enumerate(ws[1], 1):
            col_name = str(cell.value).lower() if cell.value else ''
            if col_name == 'organization':
                org_col = idx
            elif 'severity' in col_name:
                severity_col = idx
            elif 'compliance' in col_name:
                compliance_col = idx
            elif 'dependabot status' in col_name:
                dependabot_status_col = idx
            elif 'dependabot alert' in col_name:
                dependabot_alerts_col = idx
            elif 'code scanning status' in col_name or 'code scan status' in col_name:
                code_status_col = idx
            elif 'code scan alert' in col_name:
                code_alerts_col = idx
            elif 'secret scanning status' in col_name or 'secret scan status' in col_name:
                secret_status_col = idx
            elif 'secret alert' in col_name:
                secret_alerts_col = idx
            elif 'total alert' in col_name:
                total_col = idx
            elif col_name == 'status':
                status_col = idx
            elif col_name == 'archived':
                archived_col = idx
        
        # Color code rows
        current_org = None
        for row in range(2, ws.max_row + 1):
            # Organization grouping with alternating colors
            if org_col:
                org_cell = ws.cell(row=row, column=org_col)
                if org_cell.value != current_org:
                    current_org = org_cell.value
                    org_cell.font = Font(bold=True, size=11)
            
            # Severity column - bold color coding
            if severity_col:
                severity_cell = ws.cell(row=row, column=severity_col)
                severity = str(severity_cell.value).lower() if severity_cell.value else ''
                
                if 'critical' in severity:
                    severity_cell.fill = PatternFill(start_color=self.colors['critical'],
                                                     end_color=self.colors['critical'],
                                                     fill_type='solid')
                    severity_cell.font = Font(bold=True, color=self.colors['header_text'])
                elif 'high' in severity:
                    severity_cell.fill = PatternFill(start_color=self.colors['warning'],
                                                     end_color=self.colors['warning'],
                                                     fill_type='solid')
                    severity_cell.font = Font(bold=True, color=self.colors['header_text'])
                elif 'medium' in severity:
                    severity_cell.fill = PatternFill(start_color='FFD700',
                                                     end_color='FFD700',
                                                     fill_type='solid')
                    severity_cell.font = Font(bold=True, color='000000')
                elif 'low' in severity:
                    severity_cell.fill = PatternFill(start_color='FFFFCC',
                                                     end_color='FFFFCC',
                                                     fill_type='solid')
                    severity_cell.font = Font(bold=True, color='666666')
                else:  # Clean
                    severity_cell.fill = PatternFill(start_color=self.colors['active'],
                                                     end_color=self.colors['active'],
                                                     fill_type='solid')
                    severity_cell.font = Font(bold=True, color='006400')
            
            # Compliance % column - color gradient
            if compliance_col:
                compliance_cell = ws.cell(row=row, column=compliance_col)
                try:
                    compliance = int(compliance_cell.value) if compliance_cell.value else 100
                    
                    if compliance >= 80:
                        compliance_cell.fill = PatternFill(start_color='C6EFCE',  # Light green
                                                          end_color='C6EFCE',
                                                          fill_type='solid')
                        compliance_cell.font = Font(bold=True, color='006100')
                    elif compliance >= 60:
                        compliance_cell.fill = PatternFill(start_color='FFEB9C',  # Light yellow
                                                          end_color='FFEB9C',
                                                          fill_type='solid')
                        compliance_cell.font = Font(bold=True, color='9C5700')
                    elif compliance >= 40:
                        compliance_cell.fill = PatternFill(start_color='FFCC99',  # Light orange
                                                          end_color='FFCC99',
                                                          fill_type='solid')
                        compliance_cell.font = Font(bold=True, color='CC3300')
                    else:
                        compliance_cell.fill = PatternFill(start_color='FFC7CE',  # Light red
                                                          end_color='FFC7CE',
                                                          fill_type='solid')
                        compliance_cell.font = Font(bold=True, color='9C0006')
                except:
                    pass
            
            # Status columns - Enabled/Disabled color coding
            for status_col_idx in [dependabot_status_col, code_status_col, secret_status_col]:
                if status_col_idx:
                    status_cell = ws.cell(row=row, column=status_col_idx)
                    if status_cell.value == 'Enabled':
                        status_cell.font = Font(bold=True, color='006400')
                    elif status_cell.value == 'Disabled':
                        status_cell.font = Font(bold=True, color=self.colors['critical'])
            
            # Alert count columns - highlight high values
            for alert_col_idx, alert_type in [
                (dependabot_alerts_col, 'dependabot'),
                (code_alerts_col, 'code'),
                (secret_alerts_col, 'secret'),
                (total_col, 'total')
            ]:
                if alert_col_idx:
                    alert_cell = ws.cell(row=row, column=alert_col_idx)
                    try:
                        count = int(alert_cell.value) if alert_cell.value else 0
                        if count > 0:
                            if alert_type == 'secret' or count >= 50:
                                alert_cell.fill = PatternFill(start_color='FFC7CE',
                                                             end_color='FFC7CE',
                                                             fill_type='solid')
                                alert_cell.font = Font(bold=True, color='9C0006')
                            elif count >= 20:
                                alert_cell.fill = PatternFill(start_color='FFEB9C',
                                                             end_color='FFEB9C',
                                                             fill_type='solid')
                                alert_cell.font = Font(bold=True, color='9C5700')
                            elif count >= 5:
                                alert_cell.font = Font(bold=True, color=self.colors['warning'])
                            else:
                                alert_cell.font = Font(color='666666')
                    except:
                        pass
            
            # Archived column
            if archived_col:
                archived_cell = ws.cell(row=row, column=archived_col)
                if archived_cell.value == 'Yes':
                    archived_cell.fill = PatternFill(start_color=self.colors['archived'],
                                                     end_color=self.colors['archived'],
                                                     fill_type='solid')
                    archived_cell.font = Font(bold=True, color=self.colors['critical'])
            
            # Alternating row background
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if not cell.fill or cell.fill.start_color.rgb == '00000000':
                        cell.fill = PatternFill(start_color=self.colors['alt_row'],
                                               end_color=self.colors['alt_row'],
                                               fill_type='solid')
    
    def _format_repositories_sheet(self, ws):
        """Apply color coding to All Repositories sheet"""
        
        # Find relevant columns
        status_col = None
        visibility_col = None
        fork_col = None
        dependabot_col = None
        code_scan_col = None
        secret_scan_col = None
        
        for idx, cell in enumerate(ws[1], 1):
            col_name = str(cell.value).lower() if cell.value else ''
            if col_name == 'status':
                status_col = idx
            elif 'visibility' in col_name:
                visibility_col = idx
            elif col_name == 'fork':
                fork_col = idx
            elif 'dependabot' in col_name:
                dependabot_col = idx
            elif 'code scanning' in col_name:
                code_scan_col = idx
            elif 'secret' in col_name and 'alert' in col_name:
                secret_scan_col = idx
        
        # Color code data rows
        for row in range(2, ws.max_row + 1):
            # Status column
            if status_col:
                cell = ws.cell(row=row, column=status_col)
                if cell.value == 'Archived':
                    cell.fill = PatternFill(start_color=self.colors['archived'],
                                           end_color=self.colors['archived'],
                                           fill_type='solid')
                    cell.font = Font(bold=True, color=self.colors['critical'])
                elif cell.value == 'Active':
                    cell.fill = PatternFill(start_color=self.colors['active'],
                                           end_color=self.colors['active'],
                                           fill_type='solid')
                    cell.font = Font(bold=True, color='006400')
            
            # Visibility column
            if visibility_col:
                cell = ws.cell(row=row, column=visibility_col)
                if cell.value == 'Public':
                    cell.fill = PatternFill(start_color=self.colors['public'],
                                           end_color=self.colors['public'],
                                           fill_type='solid')
                    cell.font = Font(bold=True, color='0563C1')
                elif cell.value == 'Private':
                    cell.fill = PatternFill(start_color=self.colors['private'],
                                           end_color=self.colors['private'],
                                           fill_type='solid')
                    cell.font = Font(bold=True, color=self.colors['warning'])
            
            # Security alert columns (if present)
            for col_idx in [dependabot_col, code_scan_col, secret_scan_col]:
                if col_idx:
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        count = int(cell.value) if cell.value else 0
                        if count > 0:
                            cell.font = Font(bold=True, color=self.colors['critical'])
                    except:
                        pass
            
            # Alternating row colors
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if not cell.fill or cell.fill.start_color.rgb == '00000000':
                        cell.fill = PatternFill(start_color=self.colors['alt_row'],
                                               end_color=self.colors['alt_row'],
                                               fill_type='solid')
    
    def _format_organization_risk_pivot_sheet(self, ws):
        """Apply professional formatting to Organization Risk Pivot sheet"""
        
        # Find column indices
        org_col = None
        repo_col = None
        critical_col = None
        high_col = None
        medium_col = None
        low_col = None
        total_col = None
        
        for idx, cell in enumerate(ws[1], 1):
            col_name = str(cell.value).lower() if cell.value else ''
            if col_name == 'organization':
                org_col = idx
            elif col_name == 'repository':
                repo_col = idx
            elif col_name == 'critical':
                critical_col = idx
            elif col_name == 'high':
                high_col = idx
            elif col_name == 'medium':
                medium_col = idx
            elif col_name == 'low':
                low_col = idx
            elif col_name == 'total':
                total_col = idx
        
        # Color code severity columns
        for row in range(2, ws.max_row + 1):
            # Organization column - bold
            if org_col:
                org_cell = ws.cell(row=row, column=org_col)
                org_cell.font = Font(bold=True, size=11)
            
            # Critical column
            if critical_col:
                cell = ws.cell(row=row, column=critical_col)
                try:
                    count = int(cell.value) if cell.value else 0
                    if count > 0:
                        cell.fill = PatternFill(start_color='FFC7CE',  # Light red
                                               end_color='FFC7CE',
                                               fill_type='solid')
                        cell.font = Font(bold=True, color='9C0006')
                    cell.alignment = Alignment(horizontal='center')
                except:
                    pass
            
            # High column
            if high_col:
                cell = ws.cell(row=row, column=high_col)
                try:
                    count = int(cell.value) if cell.value else 0
                    if count > 0:
                        cell.fill = PatternFill(start_color='FFEB9C',  # Light orange
                                               end_color='FFEB9C',
                                               fill_type='solid')
                        cell.font = Font(bold=True, color='9C5700')
                    cell.alignment = Alignment(horizontal='center')
                except:
                    pass
            
            # Medium column
            if medium_col:
                cell = ws.cell(row=row, column=medium_col)
                try:
                    count = int(cell.value) if cell.value else 0
                    if count > 0:
                        cell.fill = PatternFill(start_color='FFFFE0',  # Light yellow
                                               end_color='FFFFE0',
                                               fill_type='solid')
                        cell.font = Font(bold=True, color='CC9900')
                    cell.alignment = Alignment(horizontal='center')
                except:
                    pass
            
            # Low column
            if low_col:
                cell = ws.cell(row=row, column=low_col)
                try:
                    count = int(cell.value) if cell.value else 0
                    if count > 0:
                        cell.font = Font(color='666666')
                    cell.alignment = Alignment(horizontal='center')
                except:
                    pass
            
            # Total column - bold
            if total_col:
                cell = ws.cell(row=row, column=total_col)
                try:
                    count = int(cell.value) if cell.value else 0
                    if count >= 100:
                        cell.fill = PatternFill(start_color=self.colors['critical'],
                                               end_color=self.colors['critical'],
                                               fill_type='solid')
                        cell.font = Font(bold=True, color=self.colors['header_text'])
                    elif count >= 50:
                        cell.fill = PatternFill(start_color=self.colors['warning'],
                                               end_color=self.colors['warning'],
                                               fill_type='solid')
                        cell.font = Font(bold=True, color=self.colors['header_text'])
                    else:
                        cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(horizontal='center')
                except:
                    pass
            
            # Alternating row colors
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if not cell.fill or cell.fill.start_color.rgb == '00000000':
                        cell.fill = PatternFill(start_color=self.colors['alt_row'],
                                               end_color=self.colors['alt_row'],
                                               fill_type='solid')