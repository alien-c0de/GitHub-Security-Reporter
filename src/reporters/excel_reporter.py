"""
Excel report generator  —  SecureWatch
═══════════════════════════════════════════════════════════════════════════════
IMPROVEMENTS in this version:
  1.  Executive Summary rebuilt with rich tool/company metadata from .env
  2.  Contact email always visible in Executive Summary (fix)
  3.  Weekly: Trend Analysis + Remediation Progress + Feature Summary
      merged into ONE "Analysis & Progress" tab — each section has a coloured
      heading followed immediately by its data table
  4.  Weekly: new "Repository Risk Pivot" tab — one row per repo,
      columns: Critical | High | Medium | Low | Total | Risk Level
  5.  DailyExcelReporter: Executive Summary added as first sheet
═══════════════════════════════════════════════════════════════════════════════
"""
from __future__ import annotations

from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

from src.reporters.base_reporter import BaseReporter
from config.settings import settings

logger = logging.getLogger(__name__)

# ─── Colour palette ───────────────────────────────────────────────────────────
_CLR = {
    'navy':    '1E2761',
    'white':   'FFFFFF',
    'blue':    '366092',     # original weekly header blue – kept for detail sheets
    'teal':    '065A82',
    'green':   '2C5F2D',
    'dkgreen': '006400',
    'red':     'DC143C',
    'dkred':   '8B0000',
    'orange':  'FF8C00',
    'gold':    'FFD700',
    'lime':    '90EE90',
    'pink':    'FFB6C1',
    'alt':     'F8F9FA',
    'lblue':   'E3F2FD',
}

_THIN   = Side(style='thin')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ─── Low-level cell helpers ───────────────────────────────────────────────────
def _fill(hex_: str) -> PatternFill:
    return PatternFill(start_color=hex_, end_color=hex_, fill_type='solid')


def _row_height(ws, row: int, h: int):
    ws.row_dimensions[row].height = h


def _merged_header(ws, row: int, text: str, bg: str,
                   ncols: int = 6, size: int = 13, height: int = 22):
    """Write a full-width merged section heading."""
    cell = ws.cell(row=row, column=1, value=text)
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    cell.font      = Font(bold=True, color=_CLR['white'], size=size)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    _row_height(ws, row, height)


def _col_header_row(ws, row: int, labels: List[str], bg: str = _CLR['navy']):
    """Write a row of bold column-header cells."""
    for c, label in enumerate(labels, start=1):
        cell           = ws.cell(row=row, column=c, value=label)
        cell.font      = Font(bold=True, color=_CLR['white'], size=11)
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _BORDER
    _row_height(ws, row, 18)


def _auto_width(ws, cap: int = 60):
    for col in ws.columns:
        best   = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(best + 2, cap)


def _sev_style(severity: str):
    """Return (bg_hex, fg_hex) for a severity level string."""
    s = str(severity).lower()
    if 'critical' in s: return _CLR['red'],    _CLR['white']
    if 'high'     in s: return _CLR['orange'], _CLR['white']
    if 'medium'   in s: return _CLR['gold'],   '000000'
    if 'low'      in s: return _CLR['lime'],   '000000'
    return None, None


def _alt_border(cell, row: int):
    """Apply alternating row fill + border (skip if cell already has a fill)."""
    cell.border = _BORDER
    if row % 2 == 0:
        rgb = ''
        try:
            rgb = cell.fill.start_color.rgb
        except Exception:
            pass
        if rgb in ('00000000', ''):
            cell.fill = _fill(_CLR['alt'])


# ═════════════════════════════════════════════════════════════════════════════
#  SHARED: Executive Summary writer
#  Called by both ExcelReporter (weekly) and DailyExcelReporter
# ═════════════════════════════════════════════════════════════════════════════
def _write_executive_summary(
    ws,
    report_type: str,
    open_vulns:  List[Dict],
    secrets_count: int = 0,
    org_info: Optional[Dict] = None,
):
    """
    Populate `ws` as a professional Executive Summary sheet.

    Parameters
    ----------
    ws            : openpyxl Worksheet (should be blank)
    report_type   : 'Weekly' | 'Daily'
    open_vulns    : open vulnerability alert dicts
    secrets_count : count of open secret-scanning alerts
    org_info      : {'name', 'owners', 'contact_email'} from GitHub API
    """
    org_info = org_info or {}

    # ── Safe settings reader ──────────────────────────────────────────────────
    # Handles three cases:
    #   1. Attribute exists and is a plain string  → use it
    #   2. Attribute exists but is a pydantic FieldInfo object (not yet added
    #      to settings.py properly) → fall back to the default
    #   3. Attribute does not exist at all         → use the default
    def _s(attr: str, default: str = '') -> str:
        val = getattr(settings, attr, None)
        if val is None:
            return default
        # pydantic v1 FieldInfo has 'default'; pydantic v2 has 'default' too
        if hasattr(val, 'default'):          # it's a FieldInfo – not yet resolved
            fb = val.default
            return str(fb) if fb is not None else default
        raw = str(val).strip()
        return raw if raw else default

    # ── Pull metadata from settings (.env values) ─────────────────────────────
    tool_title   = _s('report_title',     'GitHub Advanced Security Reporter')
    company      = _s('company_name',     '') or settings.github_org
    developed_by = _s('developed_by',     'Security Engineering Team')
    version      = _s('tool_version',     '1.0.0')
    copy_year    = _s('copyright_year',   '') or str(datetime.now().year)
    website      = _s('company_website',  '') or 'N/A'
    github_repo  = _s('tool_github_repo', '') or 'N/A'
    support_mail = _s('support_email',    '') or 'N/A'

    # Contact email: prefer live GitHub API result, then .env SUPPORT_EMAIL
    contact_email = (org_info.get('contact_email') or support_mail or 'N/A')

    # ── Severity counts ───────────────────────────────────────────────────────
    sev = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    for r in open_vulns:
        s = str(r.get('severity') or r.get('security_severity_level') or '').lower()
        if s in sev:
            sev[s] += 1
    total_vulns = sum(sev.values())

    r = 1   # row pointer

    # ══════════════════════════════════════════════════════════════════════════
    #  SECTION 1 – Banner
    # ══════════════════════════════════════════════════════════════════════════
    _merged_header(ws, r, f'🛡  {tool_title}', _CLR['navy'], ncols=4, size=15, height=28)
    r += 1
    meta = [
        ('Report Type',   f'{report_type} Security Report'),
        ('Generated On',  datetime.now().strftime('%A, %d %B %Y  —  %H:%M:%S')),
        ('Report Period', 'Daily – Last 24 hours'
                          if report_type == 'Daily'
                          else f'Week {datetime.now().isocalendar()[1]},  {datetime.now().year}'),
    ]
    for label, value in meta:
        ws.cell(r, 1, label).font = Font(bold=True, color=_CLR['navy'], size=11)
        ws.cell(r, 2, value).font = Font(size=11)
        r += 1
    r += 1

    # ══════════════════════════════════════════════════════════════════════════
    #  SECTION 2 – Organization Details
    # ══════════════════════════════════════════════════════════════════════════
    _merged_header(ws, r, '🏢  Organization Details', _CLR['green'], ncols=4, size=12)
    r += 1

    org_rows = [
        ('Organization Name',    org_info.get('name', company)),
        ('GitHub Organization',  settings.github_org),
        ('Organization Admins',  org_info.get('owners', 'N/A')),
        ('Contact Email',        contact_email),
        ('GitHub URL',           f'https://github.com/{settings.github_org}'),
    ]
    for label, value in org_rows:
        lc = ws.cell(r, 1, label)
        vc = ws.cell(r, 2, value)
        lc.font = Font(bold=True, color='444444', size=11)
        vc.font = Font(size=11)
        if str(value).startswith('http') or ('@' in str(value) and '.' in str(value)):
            vc.font = Font(size=11, color='0563C1', underline='single')
        if r % 2 == 0:
            lc.fill = _fill(_CLR['alt'])
            vc.fill = _fill(_CLR['alt'])
        r += 1
    r += 1

    # ══════════════════════════════════════════════════════════════════════════
    #  SECTION 3 – Security Summary
    # ══════════════════════════════════════════════════════════════════════════
    _merged_header(ws, r, '⚠️  Security Summary', _CLR['red'], ncols=4, size=12)
    r += 1

    _col_header_row(ws, r, ['Metric', 'Count', 'Status', 'SLA'], _CLR['navy'])
    r += 1

    sev_rows = [
        ('Total Open Vulnerabilities', total_vulns,   '⚠ Action Required' if total_vulns  > 0 else '✓ Clean',   '—'),
        ('Critical Severity',          sev['critical'],'🔴 Fix Immediately' if sev['critical'] > 0 else '✓ None', '48 hrs'),
        ('High Severity',              sev['high'],    '🟠 Fix This Week'   if sev['high']     > 0 else '✓ None', '7 days'),
        ('Medium Severity',            sev['medium'],  '🟡 Plan Fix'        if sev['medium']   > 0 else '✓ None', '30 days'),
        ('Low Severity',               sev['low'],     '🟢 Monitor'         if sev['low']      > 0 else '✓ None', '90 days'),
        ('Exposed Secrets',            secrets_count,  '🔴 Fix Immediately' if secrets_count   > 0 else '✓ None', '24 hrs'),
    ]

    sev_key_map = {
        'Critical Severity': 'critical', 'High Severity': 'high',
        'Medium Severity': 'medium',     'Low Severity': 'low',
        'Exposed Secrets': 'critical',
    }

    for metric, count, status, sla in sev_rows:
        mc = ws.cell(r, 1, metric)
        cc = ws.cell(r, 2, count)
        sc = ws.cell(r, 3, status)
        nc = ws.cell(r, 4, sla)

        mc.font = Font(bold=True, size=11)
        cc.alignment = Alignment(horizontal='center')
        sc.alignment = Alignment(horizontal='left')
        nc.font = Font(italic=True, color='555555', size=10)

        bg, fg = _sev_style(sev_key_map.get(metric, ''))
        if bg and count > 0:
            cc.fill = _fill(bg)
            cc.font = Font(bold=True, color=fg, size=11)

        if   '🔴' in status: sc.font = Font(bold=True, color=_CLR['dkred'],  size=11)
        elif '🟠' in status: sc.font = Font(bold=True, color='8B4513',        size=11)
        elif '🟡' in status: sc.font = Font(bold=True, color='8B8B00',        size=11)
        elif '✓'  in status: sc.font = Font(bold=True, color=_CLR['dkgreen'], size=11)

        for cell in (mc, cc, sc, nc):
            cell.border = _BORDER
            if r % 2 == 0:
                try:
                    if cell.fill.start_color.rgb in ('00000000', ''):
                        cell.fill = _fill(_CLR['alt'])
                except Exception:
                    pass
        r += 1

    r += 1

    # ══════════════════════════════════════════════════════════════════════════
    #  SECTION 4 – About This Tool  (moved to end per user request)
    # ══════════════════════════════════════════════════════════════════════════
    _merged_header(ws, r, '📋  About This Tool', _CLR['teal'], ncols=4, size=12)
    r += 1

    about_rows = [
        ('Tool Name',         tool_title),
        ('Company / Client',  company),
        ('Developed By',      developed_by),
        ('Version',           f'v{version}'),
        ('Copyright',         f'© {copy_year}  {company}'),
        ('Company Website',   website),
        ('GitHub Repository', github_repo),
        ('Support / Contact', support_mail),
    ]
    for label, value in about_rows:
        lc = ws.cell(r, 1, label)
        vc = ws.cell(r, 2, value)
        lc.font = Font(bold=True, color='444444', size=11)
        vc.font = Font(size=11)
        if str(value).startswith('http') or ('@' in str(value) and '.' in str(value)):
            vc.font = Font(size=11, color='0563C1', underline='single')
        if r % 2 == 0:
            lc.fill = _fill(_CLR['alt'])
            vc.fill = _fill(_CLR['alt'])
        r += 1

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 12


# ═════════════════════════════════════════════════════════════════════════════
#  WEEKLY  ExcelReporter
# ═════════════════════════════════════════════════════════════════════════════
class ExcelReporter(BaseReporter):
    """Generate weekly GitHub Advanced Security Excel reports."""

    def __init__(self):
        super().__init__()
        self.output_dir = settings.report_output_dir

    # ── Org info helper — reads from the already-collected snapshot ─────────
    @staticmethod
    def _resolve_org_info(snapshot: Dict[str, Any]) -> Dict[str, Any]:
        """
        Build org info dict from data that is ALREADY present in `snapshot`
        (collected earlier in the pipeline by OrganizationCollector /
        RepositoryHealthCollector).  No new GitHub API call is made here.

        Keys used from snapshot
        -----------------------
        snapshot['organization']          – org login / name
        snapshot['organization_data']     – full org payload (optional)
        snapshot['repository_health']     – list of repo dicts, each may carry
                                            owner_name, repository_admins,
                                            admin_emails
        """
        org_data  = snapshot.get('organization_data') or {}
        repo_health = snapshot.get('repository_health') or []

        # ── Org display name ─────────────────────────────────────────────────
        name = (
            org_data.get('name')
            or snapshot.get('organization')
            or settings.github_org
        )

        # ── Admins: collected per-repo by RepositoryHealthCollector ──────────
        # Gather unique admin names from repository_health rows
        admin_parts: list = []
        email_parts: list = []
        seen_admins: set  = set()
        seen_emails: set  = set()

        for repo in repo_health:
            # repository_admins may be a comma-separated string or a list
            raw_admins = repo.get('repository_admins') or ''
            if isinstance(raw_admins, list):
                raw_admins = ', '.join(str(a) for a in raw_admins)
            for part in raw_admins.split(','):
                part = part.strip()
                if part and part not in seen_admins:
                    seen_admins.add(part)
                    admin_parts.append(part)

            # admin_emails may similarly be a comma-separated string or list
            raw_emails = repo.get('admin_emails') or ''
            if isinstance(raw_emails, list):
                raw_emails = ', '.join(str(e) for e in raw_emails)
            for email in raw_emails.split(','):
                email = email.strip()
                if email and email not in seen_emails and '@' in email:
                    seen_emails.add(email)
                    email_parts.append(email)

        # Fall back to org-level email if no repo-level emails found
        if not email_parts:
            org_email = org_data.get('email') or org_data.get('contact_email') or ''
            if org_email and '@' in org_email:
                email_parts.append(org_email)

        return {
            'name':          name,
            'owners':        ', '.join(admin_parts) if admin_parts else 'N/A',
            'contact_email': ', '.join(email_parts) if email_parts else 'Not publicly available',
        }

    # ── Entry point ───────────────────────────────────────────────────────────
    def generate_report(
        self,
        snapshot: Dict[str, Any],
        metrics:  Dict[str, Any],
        trends:   Optional[Dict[str, Any]] = None,
    ) -> Path:
        from src.utils.validators import DataValidator

        org_name  = snapshot.get('organization', settings.github_org)
        sanitized = DataValidator.sanitize_filename(org_name)
        ts        = datetime.now().strftime('%Y%m%d_%H%M%S')
        week_num  = datetime.now().isocalendar()[1]

        filename  = self.output_dir / (f'{sanitized}_Weekly_Report_W{week_num:02d}_{ts}.xlsx')

        logger.info(f'[bright_yellow][+] Generating Excel report: {filename}[/bright_yellow]')

        # Derive org info from the snapshot that was already fetched upstream —
        # no second GitHub API call is made.
        org_info = self._resolve_org_info(snapshot)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            self._sheet_executive_summary(writer, metrics, trends, snapshot, org_info)  # Tab 1
            self._sheet_analysis_and_progress(writer, snapshot, trends)        # Tab 2 (merged)
            self._sheet_top_risks(writer, snapshot)                            # Tab 3
            self._sheet_repository_health(writer, snapshot)                    # Tab 4
            self._sheet_recommendations(writer, metrics)                       # Tab 5
            self._sheet_repository_risk_pivot(writer, snapshot)                # Tab 6 NEW
            self._sheet_dependabot_details(writer, snapshot.get('dependabot', []))
            self._sheet_code_scanning_details(writer, snapshot.get('code_scanning', []))
            self._sheet_secret_scanning_details(writer, snapshot.get('secret_scanning', []))
            self._sheet_supply_chain(writer, snapshot.get('supply_chain', []))

        self._apply_formatting(filename)
        logger.info(f'[bright_green]📂 Excel report generated: {filename}[/bright_green]')
        return filename

    # ── Tab 1 – Executive Summary ─────────────────────────────────────────────
    def _sheet_executive_summary(self, writer, metrics, trends, snapshot,
                                  org_info: Optional[Dict[str, Any]] = None):
        # org_info is always resolved ONCE in generate_report() and passed here.
        # No secondary GitHub API call is made — reuse whatever was collected.
        if org_info is None:
            org_info = {}   # fall back to empty dict; settings values fill the gaps

        open_vulns    = [r for r in
                         snapshot.get('dependabot', []) + snapshot.get('code_scanning', [])
                         if r.get('state') == 'open']
        secrets_count = sum(1 for r in snapshot.get('secret_scanning', [])
                            if r.get('state') == 'open')

        # create blank sheet then populate via shared writer
        pd.DataFrame({'_': []}).to_excel(
            writer, sheet_name='Executive Summary', index=False)
        ws = writer.sheets['Executive Summary']
        ws.delete_rows(1)

        _write_executive_summary(ws, 'Weekly', open_vulns, secrets_count, org_info)

        # append week-over-week change line
        if trends:
            tt  = trends.get('total_vulnerabilities', {})
            ch  = tt.get('absolute_change', 0)
            pct = tt.get('percent_change', 0)
            nr  = ws.max_row + 2
            ws.cell(nr, 1, 'Week-over-Week Change').font = Font(bold=True, color=_CLR['navy'])
            ws.cell(nr, 2, f'{ch:+d} ({pct:+.1f}%)').font = Font(bold=True)

    # ── Tab 2 – Analysis & Progress (merged Trend + Remediation + Feature) ────
    def _sheet_analysis_and_progress(self, writer, snapshot, trends):
        wb = writer.book
        ws = wb.create_sheet('Analysis & Progress')
        repo_health = snapshot.get('repository_health', [])
        r = 1

        # ╔══════════════════════════════════╗
        # ║  A — Trend Analysis              ║
        # ╚══════════════════════════════════╝
        _merged_header(ws, r, '📈  Trend Analysis', _CLR['navy'], ncols=7)
        r += 1

        if not trends:
            ws.cell(r, 1, 'Trend data not available — requires historical snapshots.').font = Font(italic=True)
            r += 3
        else:
            _col_header_row(ws, r,
                ['Category', 'Previous Week', 'Current Week',
                 'Change', 'Change %', 'Direction', 'Improved'], _CLR['teal'])
            r += 1
            for cat, info in trends.items():
                if isinstance(info, dict) and 'current' in info:
                    vals = [
                        cat.replace('_', ' ').title(),
                        info.get('previous', 0),
                        info.get('current', 0),
                        info.get('absolute_change', 0),
                        f"{info.get('percent_change', 0):.1f}%",
                        info.get('direction', 'unchanged'),
                        '✓' if info.get('improved') else '✗',
                    ]
                    for c, v in enumerate(vals, 1):
                        _alt_border(ws.cell(r, c, v), r)
                    r += 1
        r += 2

        # ╔══════════════════════════════════╗
        # ║  B — Remediation Progress        ║
        # ╚══════════════════════════════════╝
        _merged_header(ws, r, '🔧  Remediation Progress', _CLR['green'], ncols=7)
        r += 1

        remediation = []
        for a in snapshot.get('dependabot', []):
            if a.get('state') in ('dismissed', 'fixed'):
                remediation.append({
                    'Type': 'Dependency',
                    'Repository': a.get('repository'),
                    'Severity': str(a.get('severity', '')).title(),
                    'Description': f"{a.get('package_name','N/A')} — {str(a.get('summary',''))[:60]}",
                    'Closed Date': a.get('dismissed_at') or a.get('fixed_at', 'N/A'),
                    'Resolution': a.get('dismissed_reason', 'Fixed'),
                })
        for a in snapshot.get('code_scanning', []):
            if a.get('state') in ('dismissed', 'fixed'):
                remediation.append({
                    'Type': 'Code Scanning',
                    'Repository': a.get('repository'),
                    'Severity': str(a.get('security_severity_level', '')).title(),
                    'Description': str(a.get('rule_description', ''))[:60],
                    'Closed Date': a.get('dismissed_at') or a.get('fixed_at', 'N/A'),
                    'Resolution': a.get('dismissed_reason', 'Fixed'),
                })
        for a in snapshot.get('secret_scanning', []):
            if a.get('state') == 'resolved':
                remediation.append({
                    'Type': 'Secret',
                    'Repository': a.get('repository'),
                    'Severity': 'High',
                    'Description': str(a.get('secret_type', '')),
                    'Closed Date': a.get('resolved_at', 'N/A'),
                    'Resolution': a.get('resolution', 'Revoked'),
                })

        rem_cols = ['Type', 'Repository', 'Severity', 'Description', 'Closed Date', 'Resolution']
        _col_header_row(ws, r, rem_cols, _CLR['green'])
        r += 1

        if remediation:
            for item in remediation:
                for c, key in enumerate(rem_cols, 1):
                    cell = ws.cell(r, c, item.get(key, ''))
                    _alt_border(cell, r)
                    if key == 'Severity':
                        bg, fg = _sev_style(str(item.get('Severity', '')))
                        if bg:
                            cell.fill = _fill(bg)
                            cell.font = Font(bold=True, color=fg)
                r += 1
        else:
            ws.cell(r, 1, 'No vulnerabilities closed this period.').font = Font(italic=True)
            r += 1
        r += 2

        # ╔══════════════════════════════════╗
        # ║  C — Feature Status Summary      ║
        # ╚══════════════════════════════════╝
        _merged_header(ws, r, '🔩  Feature Status Summary', _CLR['teal'], ncols=7)
        r += 1

        if not repo_health:
            ws.cell(r, 1, 'No repository health data available.').font = Font(italic=True)
        else:
            df_h  = pd.DataFrame(repo_health)
            total = len(df_h)
            feats = [
                ('Dependabot',        'dependabot_enabled'),
                ('Code Scanning',     'code_scanning_enabled'),
                ('Secret Scanning',   'secret_scanning_enabled'),
                ('Branch Protection', 'branch_protection_enabled'),
                ('Security Policy',   'has_security_policy'),
            ]
            _col_header_row(ws, r, ['Feature', 'Enabled', 'Disabled / N/A',
                                     'Total Repos', 'Coverage %'], _CLR['teal'])
            r += 1
            for name, key in feats:
                if key not in df_h.columns:
                    continue
                enabled  = int(df_h[key].sum())
                disabled = total - enabled
                pct_val  = enabled / total * 100 if total else 0
                vals     = [name, enabled, disabled, total, f'{pct_val:.1f}%']
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(r, c, v)
                    _alt_border(cell, r)
                    if c == 5:
                        if pct_val >= 80:
                            cell.fill = _fill(_CLR['lime'])
                            cell.font = Font(bold=True, color=_CLR['dkgreen'])
                        else:
                            cell.fill = _fill(_CLR['pink'])
                            cell.font = Font(bold=True, color=_CLR['dkred'])
                r += 1

        _auto_width(ws)
        ws.freeze_panes = 'A2'

    # ── Tab 3 – Top Risks ─────────────────────────────────────────────────────
    def _sheet_top_risks(self, writer, snapshot):
        risks = []
        for a in snapshot.get('dependabot', []):
            if a.get('state') == 'open' and a.get('severity') in ('critical', 'high'):
                risks.append({
                    'Type': 'Dependency',    'Repository': a.get('repository'),
                    'Severity': a.get('severity'),       'Package': a.get('package_name'),
                    'CVE': a.get('cve_id', 'N/A'),       'Summary': str(a.get('summary', ''))[:100],
                    'Age (days)': a.get('age_days', 0),  'URL': a.get('url'),
                })
        for a in snapshot.get('code_scanning', []):
            if a.get('state') == 'open' and a.get('security_severity_level') in ('critical', 'high'):
                risks.append({
                    'Type': 'Code Scanning', 'Repository': a.get('repository'),
                    'Severity': a.get('security_severity_level'), 'Package': 'N/A',
                    'CVE': 'N/A', 'Summary': str(a.get('rule_description', ''))[:100],
                    'Age (days)': a.get('age_days', 0), 'URL': a.get('url'),
                })
        order = {'critical': 0, 'high': 1}
        risks.sort(key=lambda x: (order.get(str(x['Severity']).lower(), 2), -x['Age (days)']))
        (pd.DataFrame(risks[:50]) if risks
         else pd.DataFrame({'Message': ['No critical/high vulnerabilities! 🎉']})
         ).to_excel(writer, sheet_name='Top Risks', index=False)

    # ── Tab 4 – Repository Health ─────────────────────────────────────────────
    def _sheet_repository_health(self, writer, snapshot):
        rh = snapshot.get('repository_health', [])
        if not rh:
            pd.DataFrame({'Message': ['No repository health data.']}).to_excel(
                writer, sheet_name='Repository Health', index=False)
            return
        df   = pd.DataFrame(rh)
        cols = [
            'repository', 'owner_name', 'repository_admins', 'admin_emails',
            'compliance_percentage',
            'dependabot_status', 'dependabot_alert_count',
            'code_scanning_status', 'code_scanning_alert_count',
            'secret_scanning_status', 'secret_scanning_alert_count',
            'branch_protection_enabled', 'has_security_policy',
            'visibility', 'language', 'days_since_last_push', 'archived',
        ]
        avail = [c for c in cols if c in df.columns]
        rename = {
            'repository': 'Repository', 'owner_name': 'Owner',
            'repository_admins': 'Admins', 'admin_emails': 'Admin Emails',
            'compliance_percentage': 'Compliance %',
            'dependabot_status': 'Dependabot Status',
            'dependabot_alert_count': 'Dependabot Alerts',
            'code_scanning_status': 'Code Scanning Status',
            'code_scanning_alert_count': 'Code Scan Alerts',
            'secret_scanning_status': 'Secret Scanning Status',
            'secret_scanning_alert_count': 'Secret Alerts',
            'branch_protection_enabled': 'Branch Protection',
            'has_security_policy': 'Security Policy',
            'visibility': 'Visibility', 'language': 'Language',
            'days_since_last_push': 'Days Since Push', 'archived': 'Archived',
        }
        dfd = df[avail].rename(columns=rename)
        if 'Compliance %' in dfd.columns:
            dfd = dfd.sort_values('Compliance %', ascending=False)
        # Add Total Alerts (sum of all alert counts) before Visibility
        for alert_col in ('Dependabot Alerts', 'Code Scan Alerts', 'Secret Alerts'):
            if alert_col not in dfd.columns:
                dfd[alert_col] = 0
        dfd['Total Alerts'] = (
            dfd['Dependabot Alerts'].fillna(0).astype(int)
            + dfd['Code Scan Alerts'].fillna(0).astype(int)
            + dfd['Secret Alerts'].fillna(0).astype(int)
        )
        vis_pos = list(dfd.columns).index('Visibility')
        cols_ordered = list(dfd.columns)
        cols_ordered.remove('Total Alerts')
        cols_ordered.insert(vis_pos, 'Total Alerts')
        dfd = dfd[cols_ordered]
        dfd.to_excel(writer, sheet_name='Repository Health', index=False)
        ws = writer.sheets['Repository Health']

        # ── Build column-index lookup once ──────────────────────────────────
        col_idx = {name: i for i, name in enumerate(dfd.columns, 1)}
        nrows   = len(dfd) + 2   # +1 header +1 for range()

        # ── Compliance % — 3-tier traffic-light ─────────────────────────────
        if 'Compliance %' in col_idx:
            ci = col_idx['Compliance %']
            for ri in range(2, nrows):
                cell = ws.cell(ri, ci)
                try:
                    val = float(cell.value) if cell.value is not None else None
                except (TypeError, ValueError):
                    val = None
                if val is None:
                    continue
                cell.font = Font(bold=True)
                if val >= 80:
                    cell.fill = _fill('90EE90'); cell.font = Font(bold=True, color='006400')
                elif val >= 40:
                    cell.fill = _fill('FFFACD'); cell.font = Font(bold=True, color='8B8B00')
                else:
                    cell.fill = _fill('FFB6C1'); cell.font = Font(bold=True, color='8B0000')

        # ── Status columns (Dependabot / Code Scanning / Secret Scanning) ───
        for col_name in ('Dependabot Status', 'Code Scanning Status', 'Secret Scanning Status'):
            if col_name not in col_idx:
                continue
            ci = col_idx[col_name]
            for ri in range(2, nrows):
                cell = ws.cell(ri, ci)
                v    = str(cell.value or '').strip()
                cell.font = Font(bold=True)
                if v == 'Enabled':
                    cell.fill = _fill('90EE90'); cell.font = Font(bold=True, color='006400')
                elif v in ('Disabled', 'Not Available', 'No Access', 'No Scans'):
                    cell.fill = _fill('FFB6C1'); cell.font = Font(bold=True, color='8B0000')
                elif v == 'Unavailable':
                    cell.fill = _fill('D3D3D3'); cell.font = Font(bold=True, color='696969')
                elif 'Enabled' in v:   # e.g. "Enabled (No Scans)"
                    cell.fill = _fill('FFFACD'); cell.font = Font(bold=True, color='8B8B00')

        # ── Alert count columns — numeric threshold colour ────────────────────
        for col_name in ('Dependabot Alerts', 'Code Scan Alerts', 'Secret Alerts', 'Total Alerts'):
            if col_name not in col_idx:
                continue
            ci = col_idx[col_name]
            for ri in range(2, nrows):
                cell = ws.cell(ri, ci)
                try:
                    val = int(cell.value) if cell.value is not None else None
                except (TypeError, ValueError):
                    val = None
                if val is None:
                    continue
                cell.alignment = Alignment(horizontal='center')
                if col_name == 'Total Alerts':
                    if val == 0:    cell.fill = _fill('90EE90'); cell.font = Font(bold=True, color='006400')
                    elif val <= 10: cell.fill = _fill('FFFACD'); cell.font = Font(bold=True, color='8B8B00')
                    else:           cell.fill = _fill('FFB6C1'); cell.font = Font(bold=True, color='8B0000')
                else:
                    if val == 0:   cell.fill = _fill('90EE90'); cell.font = Font(bold=True, color='006400')
                    elif val <= 5: cell.fill = _fill('FFFACD'); cell.font = Font(bold=True, color='8B8B00')
                    else:          cell.fill = _fill('FFB6C1'); cell.font = Font(bold=True, color='8B0000')

        # ── Branch Protection — True=green, False=red ─────────────────────────
        if 'Branch Protection' in col_idx:
            ci = col_idx['Branch Protection']
            for ri in range(2, nrows):
                cell = ws.cell(ri, ci)
                v    = cell.value
                cell.font = Font(bold=True)
                if v in (True, 'True', 'Yes', 'yes', 1):
                    cell.fill = _fill('90EE90'); cell.font = Font(bold=True, color='006400')
                elif v in (False, 'False', 'No', 'no', 0):
                    cell.fill = _fill('FFB6C1'); cell.font = Font(bold=True, color='8B0000')

        # ── Security Policy — True=green, False=red ───────────────────────────
        if 'Security Policy' in col_idx:
            ci = col_idx['Security Policy']
            for ri in range(2, nrows):
                cell = ws.cell(ri, ci)
                v    = cell.value
                if v in (True, 'True', 'Yes', 'yes', 1):
                    cell.fill = _fill('90EE90'); cell.font = Font(bold=True, color='006400')
                elif v in (False, 'False', 'No', 'no', 0):
                    cell.fill = _fill('FFB6C1'); cell.font = Font(bold=True, color='8B0000')

    # ── Tab 5 – Recommendations ───────────────────────────────────────────────
    def _sheet_recommendations(self, writer, metrics):
        s   = metrics.get('summary', {})
        rec = []
        if s.get('critical_count', 0) > 0:
            rec.append({'Priority': '🔴 URGENT', 'Area': 'Critical Vulnerabilities',
                        'Recommendation': f"Address {s['critical_count']} critical vulnerabilities immediately.",
                        'Owner': 'Security Team + Dev Teams', 'Due Date': 'Within 48 hours'})
        if s.get('exposed_secrets', 0) > 0:
            rec.append({'Priority': '🔴 URGENT', 'Area': 'Secret Exposure',
                        'Recommendation': f"Revoke and rotate {s['exposed_secrets']} exposed secrets.",
                        'Owner': 'Security Team', 'Due Date': 'Within 24 hours'})
        if s.get('high_count', 0) > 10:
            rec.append({'Priority': '🟠 HIGH', 'Area': 'High Severity Vulnerabilities',
                        'Recommendation': f"Create remediation plan for {s['high_count']} high issues.",
                        'Owner': 'Dev Team Leads', 'Due Date': 'Within 7 days'})
        for feat, d in metrics.get('repository_health', {}).get('security_features', {}).items():
            if d.get('coverage_percentage', 0) < 80:
                rec.append({'Priority': '🟡 MEDIUM', 'Area': feat.replace('_', ' ').title(),
                            'Recommendation': f"Enable on {d.get('total',0)-d.get('enabled',0)} repos to reach 80%.",
                            'Owner': 'Platform Team', 'Due Date': 'Within 14 days'})
        if s.get('vulnerabilities_closed_this_week', 0) > 5:
            rec.append({'Priority': '🟢 INFO', 'Area': 'Team Recognition',
                        'Recommendation': f"Great work! {s['vulnerabilities_closed_this_week']} vulnerabilities resolved.",
                        'Owner': 'All Teams', 'Due Date': 'N/A'})
        if not rec:
            rec.append({'Priority': '🟢 INFO', 'Area': 'General',
                        'Recommendation': 'Security posture is strong. Continue monitoring.',
                        'Owner': 'Security Team', 'Due Date': 'Ongoing'})
        df = pd.DataFrame(rec)
        df.to_excel(writer, sheet_name='Recommendations', index=False)
        ws = writer.sheets['Recommendations']
        cmap = {
            '🔴 URGENT': ('FF0000', 'FFFFFF'), '🟠 HIGH':   ('FFA500', 'FFFFFF'),
            '🟡 MEDIUM': ('FFFF00', '000000'), '🟢 INFO':   ('90EE90', '000000'),
        }
        for row in range(2, len(df) + 2):
            cell      = ws[f'A{row}']
            bg, fg    = cmap.get(cell.value, (None, None))
            if bg:
                cell.fill = _fill(bg)
                cell.font = Font(color=fg, bold=True)

    # ── Tab 6 – Repository Risk Pivot  (NEW) ──────────────────────────────────
    def _sheet_repository_risk_pivot(self, writer, snapshot):
        """
        One row per repository.
        Columns: Repository | Critical | High | Medium | Low | Total | Risk Level
        Cells are colour-coded; a grand-totals footer is added at the bottom.
        """
        wb = writer.book
        ws = wb.create_sheet('Repository Risk Pivot')

        _merged_header(ws, 1,
            '📊  Repository Risk Pivot — Open Vulnerabilities by Severity',
            _CLR['navy'], ncols=7, size=13, height=24)

        _col_header_row(ws, 2,
            ['Repository', 'Critical', 'High', 'Medium', 'Low', 'Total', 'Risk Level'],
            _CLR['navy'])

        # Aggregate open Dependabot + Code-Scanning alerts per repo
        counts: Dict[str, Dict[str, int]] = {}
        def _add(repo, sev):
            if not repo:
                return
            counts.setdefault(repo, {'critical': 0, 'high': 0, 'medium': 0, 'low': 0})
            s = str(sev).lower()
            if s in counts[repo]:
                counts[repo][s] += 1

        for a in snapshot.get('dependabot', []):
            if a.get('state') == 'open':
                _add(a.get('repository'), a.get('severity', ''))
        for a in snapshot.get('code_scanning', []):
            if a.get('state') == 'open':
                _add(a.get('repository'), a.get('security_severity_level', ''))

        # Sort: most critical first, then high, medium, low
        sorted_repos = sorted(
            counts.items(),
            key=lambda x: (-x[1]['critical'], -x[1]['high'], -x[1]['medium'], -x[1]['low']),
        )

        risk_styles = {
            'critical': ('🔴 Critical', 'FFD7D7', _CLR['dkred']),
            'high':     ('🟠 High',     'FFE4B5', '8B4513'),
            'medium':   ('🟡 Medium',   'FFFFF0', '8B8B00'),
            'low':      ('🟢 Low',      'F0FFF0', _CLR['dkgreen']),
        }

        data_row = 3
        for repo, c in sorted_repos:
            cr, hi, me, lo = c['critical'], c['high'], c['medium'], c['low']
            total = cr + hi + me + lo

            if cr > 0:   rk = 'critical'
            elif hi > 0: rk = 'high'
            elif me > 0: rk = 'medium'
            else:        rk = 'low'
            risk_label, risk_bg, risk_fg = risk_styles[rk]

            sev_map = {2: (cr, 'critical'), 3: (hi, 'high'), 4: (me, 'medium'), 5: (lo, 'low')}

            for col, val in enumerate([repo, cr, hi, me, lo, total, risk_label], 1):
                cell           = ws.cell(data_row, col, val)
                cell.border    = _BORDER
                cell.alignment = Alignment(horizontal='left' if col == 1 else 'center')
                if data_row % 2 == 0:
                    cell.fill = _fill(_CLR['alt'])
                # colour severity count cells
                if col in sev_map:
                    cnt, sk = sev_map[col]
                    bg, fg  = _sev_style(sk)
                    if bg and cnt > 0:
                        cell.fill = _fill(bg)
                        cell.font = Font(bold=True, color=fg)
                # colour risk level cell
                if col == 7:
                    cell.fill = _fill(risk_bg)
                    cell.font = Font(bold=True, color=risk_fg)

            ws.cell(data_row, 1).font = Font(bold=True)
            data_row += 1

        # Grand-totals footer
        if sorted_repos:
            tc = sum(v['critical'] for _, v in sorted_repos)
            th = sum(v['high']     for _, v in sorted_repos)
            tm = sum(v['medium']   for _, v in sorted_repos)
            tl = sum(v['low']      for _, v in sorted_repos)
            tg = tc + th + tm + tl
            for col, val in enumerate(['TOTAL', tc, th, tm, tl, tg, ''], 1):
                cell           = ws.cell(data_row, col, val)
                cell.font      = Font(bold=True, color=_CLR['white'])
                cell.fill      = _fill(_CLR['navy'])
                cell.border    = _BORDER
                cell.alignment = Alignment(horizontal='left' if col == 1 else 'center')

        ws.freeze_panes = 'A3'
        ws.column_dimensions['A'].width = 36
        for ltr in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[ltr].width = 12
        ws.column_dimensions['G'].width = 18

    # ── Detail sheets ─────────────────────────────────────────────────────────
    def _sheet_dependabot_details(self, writer, data):
        if not data:
            pd.DataFrame({'Message': ['No Dependabot data.']}).to_excel(
                writer, sheet_name='Dependabot Details', index=False); return
        df   = pd.DataFrame(data)
        cols = ['repository', 'state', 'severity', 'package_name', 'package_ecosystem',
                'cve_id', 'summary', 'age_days', 'url']
        df[[c for c in cols if c in df.columns]].to_excel(
            writer, sheet_name='Dependabot Details', index=False)

    def _sheet_code_scanning_details(self, writer, data):
        if not data:
            pd.DataFrame({'Message': ['No Code Scanning data.']}).to_excel(
                writer, sheet_name='Code Scanning Details', index=False); return
        df   = pd.DataFrame(data)
        cols = ['repository', 'state', 'security_severity_level', 'rule_description',
                'tool_name', 'file_path', 'age_days', 'url']
        df[[c for c in cols if c in df.columns]].to_excel(
            writer, sheet_name='Code Scanning Details', index=False)

    def _sheet_secret_scanning_details(self, writer, data):
        if not data:
            pd.DataFrame({'Message': ['No Secret Scanning data.']}).to_excel(
                writer, sheet_name='Secret Scanning Details', index=False); return
        df   = pd.DataFrame(data)
        cols = ['repository', 'state', 'secret_type', 'resolution',
                'push_protection_bypassed', 'age_days', 'url']
        df[[c for c in cols if c in df.columns]].to_excel(
            writer, sheet_name='Secret Scanning Details', index=False)

    # ── Supply Chain sheet ───────────────────────────────────────────────────
    def _sheet_supply_chain(self, writer, data):
        """Supply Chain sheet with Yes/No colour coding matching the old report."""
        if not data:
            pd.DataFrame({'Message': ['No Supply Chain data available.']}).to_excel(
                writer, sheet_name='Supply Chain', index=False)
            return

        df   = pd.DataFrame(data)
        cols = [
            'repository', 'dependency_review_enabled', 'has_dependency_files',
            'dependency_graph_enabled', 'total_dependencies',
            'language', 'visibility',
        ]
        avail = [c for c in cols if c in df.columns]
        dfd   = df[avail].rename(columns={
            'repository':                 'Repository',
            'dependency_review_enabled':  'Dependency Review',
            'has_dependency_files':       'Has Dependency Files',
            'dependency_graph_enabled':   'Dependency Graph',
            'total_dependencies':         'Total Dependencies',
            'language':                   'Primary Language',
            'visibility':                 'Visibility',
        })

        # Convert booleans → Yes / No strings for display
        bool_cols = ['Dependency Review', 'Has Dependency Files', 'Dependency Graph']
        for col in bool_cols:
            if col in dfd.columns:
                dfd[col] = dfd[col].apply(
                    lambda x: 'Yes' if x in (True, 'True', 'yes', 1) else 'No'
                )

        dfd = dfd.sort_values('Repository')
        dfd.to_excel(writer, sheet_name='Supply Chain', index=False)

        ws = writer.sheets['Supply Chain']

        # Colour Yes/No cells — exact colours from old report:
        #   Yes → bg=90EE90 (lime green),  fg=006400 (dark green), bold
        #   No  → bg=FFB6C1 (light pink),  fg=8B0000 (dark red),   bold
        yes_fill = _fill('90EE90')
        no_fill  = _fill('FFB6C1')
        yes_font = Font(bold=True, color='006400')
        no_font  = Font(bold=True, color='8B0000')

        for ci, col_name in enumerate(dfd.columns, 1):
            if col_name not in bool_cols:
                continue
            for ri in range(2, len(dfd) + 2):
                cell = ws.cell(ri, ci)
                if cell.value == 'Yes':
                    cell.fill = yes_fill
                    cell.font = yes_font
                elif cell.value == 'No':
                    cell.fill = no_fill
                    cell.font = no_font

    # ── Global formatting pass ─────────────────────────────────────────────────
    def _apply_formatting(self, filename: Path):
        wb    = load_workbook(filename)
        h_fil = _fill(_CLR['blue'])
        h_fnt = Font(color='FFFFFF', bold=True, size=11)
        skip  = {'Executive Summary', 'Analysis & Progress', 'Repository Risk Pivot'}

        for sn in wb.sheetnames:
            ws = wb[sn]
            _auto_width(ws)
            if sn in skip or ws.max_row < 1:
                continue
            for cell in ws[1]:
                cell.fill      = h_fil
                cell.font      = h_fnt
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(filename)


# ═════════════════════════════════════════════════════════════════════════════
#  DAILY  DailyExcelReporter
# ═════════════════════════════════════════════════════════════════════════════
class DailyExcelReporter:
    """Enhanced daily reporter — now includes Executive Summary as first tab."""

    def __init__(self):
        self.output_dir = settings.report_output_dir
        self.colors = {
            'header_bg':   '1E2761',
            'header_text': 'FFFFFF',
            'critical':    'DC143C',
            'high':        'FF8C00',
            'medium':      'FFD700',
            'low':         '90EE90',
            'alt_row':     'F8F9FA',
        }

    # ── Org info helper — reads from already-collected snapshot (no API call) ──
    @staticmethod
    def _resolve_org_info(snapshot: Dict[str, Any]) -> Dict[str, Any]:
        """
        Reuses ExcelReporter._resolve_org_info logic: extracts org name,
        admins and contact email from data already present in `snapshot`.
        No new GitHub API call is made — the data was collected upstream.
        """
        # Delegate to the shared implementation on ExcelReporter
        return ExcelReporter._resolve_org_info(snapshot)

    def generate_daily_report(self, summary: Dict[str, Any]) -> Path:
        from src.utils.validators import DataValidator
        org_name  = getattr(settings, 'github_org', 'SecureWatch')
        sanitized = DataValidator.sanitize_filename(org_name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename  = self.output_dir / f'{sanitized}_Daily_Report_{timestamp}.xlsx'
        logger.info(f'[cyan]Generating daily report: {filename}[/cyan]')

        # Extract org info from the summary dict that was already built upstream —
        # no new GitHub API call; summary carries the same keys as snapshot.
        org_info = self._resolve_org_info(summary)

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Tab 1: Executive Summary — org_info passed in, no extra API call
            self._sheet_executive_summary(writer, summary, org_info)
            # Tab 2: Critical Items
            if summary.get('critical_items'):
                self._create_critical_items_sheet(writer, summary['critical_items'])
            # Tab 3: Exposed Secrets
            if summary.get('exposed_secrets'):
                self._create_exposed_secrets_sheet(writer, summary['exposed_secrets'])
            # Tab 4: Summary Dashboard
            self._create_summary_dashboard(writer, summary)

        self._apply_formatting(filename, summary)
        logger.info('[bright_green]✓ Daily report generated successfully[/bright_green]')
        return filename

    # ── Tab 1: Daily Executive Summary ───────────────────────────────────────
    def _sheet_executive_summary(
        self,
        writer,
        summary:  Dict[str, Any],
        org_info: Optional[Dict[str, Any]] = None,
    ):
        """
        Daily Executive Summary — identical layout to the weekly report.

        Calls the shared _write_executive_summary() function (Sections 1-4:
        Banner, Organisation Details, Security Summary, About This Tool) and
        then appends a daily-specific Section 5: New Alerts breakdown by
        type and severity for the last 24 hours.
        """
        # ── Build open-vuln list from daily critical_items ─────────────────
        # critical_items uses 'severity' key; _write_executive_summary also
        # checks 'security_severity_level', so either shape is handled.
        open_vulns    = summary.get('critical_items', [])
        secrets_count = len(summary.get('exposed_secrets', []))

        # ── Create blank sheet and populate via the shared weekly writer ───
        pd.DataFrame({'_': []}).to_excel(
            writer, sheet_name='Executive Summary', index=False)
        ws = writer.sheets['Executive Summary']
        ws.delete_rows(1)

        # Sections 1–4 written exactly as in the weekly report
        _write_executive_summary(ws, 'Daily', open_vulns, secrets_count, org_info)

        # ══════════════════════════════════════════════════════════════════
        #  SECTION 5 – Daily New-Alert Breakdown  (daily-only addition)
        # ══════════════════════════════════════════════════════════════════
        # Count by severity and type from today's critical_items
        sev_counts  = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
        type_counts: Dict[str, int] = {}
        for item in open_vulns:
            s = str(item.get('severity') or item.get('security_severity_level') or '').lower()
            if s in sev_counts:
                sev_counts[s] += 1
            t = item.get('type', 'Unknown')
            type_counts[t] = type_counts.get(t, 0) + 1
        total_today = sum(sev_counts.values())

        r = ws.max_row + 2   # start below the shared content

        _merged_header(ws, r,
            '📅  New Alerts — Last 24 Hours',
            _CLR['teal'], ncols=4, size=12)
        r += 1
        _col_header_row(ws, r,
            ['Metric', 'Count', 'Status', 'Action Required'],
            _CLR['navy'])
        r += 1

        daily_rows = [
            # (label, count, sev_key, status, action)
            ('Total New Alerts Today',
             total_today, '',
             '⚠ Review Now'       if total_today          > 0 else '✓ None',
             'Triage all new alerts'),
            ('  ↳ Critical Severity',
             sev_counts['critical'], 'critical',
             '🔴 Fix Immediately' if sev_counts['critical'] > 0 else '✓ None',
             'Patch within 48 hrs'),
            ('  ↳ High Severity',
             sev_counts['high'], 'high',
             '🟠 Fix This Week'   if sev_counts['high']     > 0 else '✓ None',
             'Patch within 7 days'),
            ('  ↳ Medium Severity',
             sev_counts['medium'], 'medium',
             '🟡 Plan Fix'        if sev_counts['medium']   > 0 else '✓ None',
             'Patch within 30 days'),
            ('  ↳ Low Severity',
             sev_counts['low'], 'low',
             '🟢 Monitor'         if sev_counts['low']      > 0 else '✓ None',
             'Patch within 90 days'),
            ('Exposed Secrets',
             secrets_count, 'critical',
             '🔴 Fix Immediately' if secrets_count          > 0 else '✓ None',
             'Revoke & rotate NOW'),
        ]
        # Append per-alert-type breakdown rows
        for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
            daily_rows.append((f'  ↳ {t}', cnt, '', '', 'See detail tab'))

        for label, count, sev_key, status, action in daily_rows:
            mc = ws.cell(r, 1, label)
            cc = ws.cell(r, 2, count)
            sc = ws.cell(r, 3, status)
            nc = ws.cell(r, 4, action)

            mc.font      = Font(bold=not label.startswith('  '), size=11)
            cc.alignment = Alignment(horizontal='center')
            nc.font      = Font(italic=True, color='555555', size=10)

            bg, fg = _sev_style(sev_key)
            if bg and count > 0:
                cc.fill = _fill(bg)
                cc.font = Font(bold=True, color=fg, size=11)

            if   '🔴' in status: sc.font = Font(bold=True, color=_CLR['dkred'],  size=11)
            elif '🟠' in status: sc.font = Font(bold=True, color='8B4513',        size=11)
            elif '🟡' in status: sc.font = Font(bold=True, color='8B8B00',        size=11)
            elif '✓'  in status: sc.font = Font(bold=True, color=_CLR['dkgreen'], size=11)

            for cell in (mc, cc, sc, nc):
                cell.border = _BORDER
                if r % 2 == 0:
                    try:
                        if cell.fill.start_color.rgb in ('00000000', ''):
                            cell.fill = _fill(_CLR['alt'])
                    except Exception:
                        pass
            r += 1

        # Column widths match the weekly Executive Summary (set by _write_executive_summary)
        # Override column D to be wider for the longer "Action Required" text
        ws.column_dimensions['D'].width = 28

    # ── Existing daily sheet methods (unchanged) ──────────────────────────────
    def _create_critical_items_sheet(self, writer, critical_items):
        df = pd.DataFrame(critical_items)
        col_order = [c for c in ['type', 'repository', 'severity', 'description']
                     if c in df.columns]
        df = df[col_order]
        df.columns = ['Alert Type', 'Repository', 'Severity', 'Description']
        sev_order  = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        df['_s']   = df['Severity'].map(sev_order)
        df         = df.sort_values('_s').drop('_s', axis=1)
        df.to_excel(writer, sheet_name='Critical Items', index=False)

    def _create_exposed_secrets_sheet(self, writer, exposed_secrets):
        df = pd.DataFrame(exposed_secrets).rename(columns={
            'repository': 'Repository', 'secret_type': 'Secret Type', 'age_days': 'Age (Days)'})
        if 'Age (Days)' in df.columns:
            df = df.sort_values('Age (Days)', ascending=False)
        df.to_excel(writer, sheet_name='Exposed Secrets', index=False)

    def _create_summary_dashboard(self, writer, summary):
        critical_count = len(summary.get('critical_items', []))
        secrets_count  = len(summary.get('exposed_secrets', []))
        type_counts    = {}
        severity_counts = {}
        for item in summary.get('critical_items', []):
            t = item.get('type', 'Unknown')
            s = item.get('severity', 'unknown')
            type_counts[t]    = type_counts.get(t, 0) + 1
            severity_counts[s] = severity_counts.get(s, 0) + 1

        data = {
            'Metric': ['Report Date', 'Total Critical Items', 'Exposed Secrets',
                       '', 'By Alert Type:'],
            'Value':  [summary.get('date', datetime.now().strftime('%Y-%m-%d')),
                       critical_count, secrets_count, '', ''],
        }
        for alert_type, count in sorted(type_counts.items()):
            data['Metric'].append(f'  • {alert_type}')
            data['Value'].append(count)
        data['Metric'].append('')
        data['Value'].append('')
        data['Metric'].append('By Severity:')
        data['Value'].append('')
        for sev, count in sorted(severity_counts.items()):
            data['Metric'].append(f'  • {sev.title()}')
            data['Value'].append(count)
        pd.DataFrame(data).to_excel(writer, sheet_name='Summary', index=False)

    # ── Formatting / Pivot ───────────────────────────────────────────────────
    def _apply_formatting(self, filename: Path, summary: Dict[str, Any]):
        wb = load_workbook(filename)

        # ── Ensure Executive Summary is the first (leftmost) tab ─────────────
        if 'Executive Summary' in wb.sheetnames:
            wb.move_sheet('Executive Summary', offset=-wb.sheetnames.index('Executive Summary'))

        # ── Format individual sheets — skip Executive Summary (already styled) ─
        if 'Critical Items' in wb.sheetnames:
            self._format_critical_items_sheet(wb['Critical Items'])
            if summary.get('critical_items'):
                self._create_pivot_analysis(wb, summary['critical_items'])
        if 'Exposed Secrets' in wb.sheetnames:
            self._format_exposed_secrets_sheet(wb['Exposed Secrets'])
        if 'Summary' in wb.sheetnames:
            self._format_summary_sheet(wb['Summary'])

        # ── Auto-width all sheets except Executive Summary ────────────────────
        # Executive Summary column widths are already set by _write_executive_summary
        # and _sheet_executive_summary — do not touch them here.
        skip = {'Executive Summary'}
        for sn in wb.sheetnames:
            if sn not in skip:
                _auto_width(wb[sn])

        wb.save(filename)

    def _format_critical_items_sheet(self, worksheet):
        hdr_fill = PatternFill(start_color=self.colors['header_bg'],
                               end_color=self.colors['header_bg'], fill_type='solid')
        hdr_font = Font(bold=True, color=self.colors['header_text'], size=11)
        for cell in worksheet[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        sev_col_idx = None
        for ci, cell in enumerate(worksheet[1], 1):
            if str(cell.value).lower() == 'severity':
                sev_col_idx = ci
                break

        sev_map = {
            'critical': (self.colors['critical'], 'FFFFFF'),
            'high':     (self.colors['high'],     'FFFFFF'),
            'medium':   (self.colors['medium'],   '000000'),
            'low':      (self.colors['low'],      '000000'),
        }
        for row in range(2, worksheet.max_row + 1):
            if sev_col_idx:
                cell = worksheet.cell(row=row, column=sev_col_idx)
                bg, fg = sev_map.get(str(cell.value).lower(), (None, None))
                if bg:
                    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                    cell.font = Font(bold=True, color=fg)
            if row % 2 == 0:
                for ci in range(1, worksheet.max_column + 1):
                    c = worksheet.cell(row=row, column=ci)
                    try:
                        if c.fill.start_color.rgb in ('00000000', ''):
                            c.fill = PatternFill(start_color=self.colors['alt_row'],
                                                 end_color=self.colors['alt_row'],
                                                 fill_type='solid')
                    except Exception:
                        pass

        for column in worksheet.columns:
            best   = max((len(str(c.value)) for c in column if c.value), default=0)
            letter = get_column_letter(column[0].column)
            worksheet.column_dimensions[letter].width = min(best + 2, 60)

    def _format_exposed_secrets_sheet(self, worksheet):
        hdr_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        for cell in worksheet[1]:
            cell.fill      = hdr_fill
            cell.font      = Font(bold=True, color='FFFFFF', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in range(2, worksheet.max_row + 1):
            age_cell = worksheet.cell(row=row, column=3)
            try:
                if age_cell.value and int(age_cell.value) > 30:
                    age_cell.fill = PatternFill(start_color='FF6B6B',
                                                end_color='FF6B6B', fill_type='solid')
                    age_cell.font = Font(bold=True, color='FFFFFF')
            except Exception:
                pass
        for column in worksheet.columns:
            best   = max((len(str(c.value)) for c in column if c.value), default=0)
            letter = get_column_letter(column[0].column)
            worksheet.column_dimensions[letter].width = min(best + 2, 60)

    def _format_summary_sheet(self, worksheet):
        for row in range(1, worksheet.max_row + 1):
            ca = worksheet.cell(row=row, column=1)
            cb = worksheet.cell(row=row, column=2)
            if ca.value and str(ca.value).endswith(':'):
                ca.font = Font(bold=True, size=12, color='1E2761')
            if 'Total' in str(ca.value):
                for c in (ca, cb):
                    c.font = Font(bold=True, size=12)
                    c.fill = PatternFill(start_color='E8F4F8',
                                         end_color='E8F4F8', fill_type='solid')
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 15

    def _create_pivot_analysis(self, workbook, critical_items):
        if 'Pivot Analysis' in workbook.sheetnames:
            del workbook['Pivot Analysis']
        ws = workbook.create_sheet('Pivot Analysis')

        ws['A1'] = 'Critical Items Analysis'
        ws['A1'].font = Font(bold=True, size=16, color='1E2761')
        ws.merge_cells('A1:D1')

        repo_counts = {}; type_counts = {}; severity_counts = {}
        for item in critical_items:
            r2 = item.get('repository', 'Unknown')
            t  = item.get('type',       'Unknown')
            s  = item.get('severity',   'unknown')
            repo_counts[r2] = repo_counts.get(r2, 0) + 1
            type_counts[t]  = type_counts.get(t,  0) + 1
            severity_counts[s] = severity_counts.get(s, 0) + 1

        cr = 3
        for section_title, data_dict in [
            ('Critical Items by Repository', repo_counts),
            ('Critical Items by Type',       type_counts),
        ]:
            ws.cell(cr, 1, section_title).font = Font(bold=True, size=12)
            cr += 1
            for col, txt in [(1, 'Category'), (2, 'Count')]:
                c = ws.cell(cr, col, txt)
                c.fill = PatternFill(start_color='1E2761', end_color='1E2761', fill_type='solid')
                c.font = Font(bold=True, color='FFFFFF')
            cr += 1
            for key, cnt in sorted(data_dict.items(), key=lambda x: -x[1]):
                ws.cell(cr, 1, key)
                ws.cell(cr, 2, cnt)
                cr += 1
            cr += 2

        # Severity section with colours
        ws.cell(cr, 1, 'Critical Items by Severity').font = Font(bold=True, size=12)
        cr += 1
        for col, txt in [(1, 'Severity'), (2, 'Count')]:
            c = ws.cell(cr, col, txt)
            c.fill = PatternFill(start_color='1E2761', end_color='1E2761', fill_type='solid')
            c.font = Font(bold=True, color='FFFFFF')
        cr += 1
        sev_colours = {
            'critical': ('DC143C', 'FFFFFF'), 'high':   ('FF8C00', 'FFFFFF'),
            'medium':   ('FFD700', '000000'), 'low':    ('90EE90', '000000'),
        }
        for sev in ('critical', 'high', 'medium', 'low'):
            if sev in severity_counts:
                bg, fg = sev_colours[sev]
                ca = ws.cell(cr, 1, sev.title())
                cb = ws.cell(cr, 2, severity_counts[sev])
                for c in (ca, cb):
                    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                    c.font = Font(bold=True, color=fg)
                cr += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15